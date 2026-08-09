from fastapi import FastAPI, Depends, HTTPException, status, Response, Request, BackgroundTasks, Header
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, cast, Date
from sqlalchemy.engine import make_url
from database import engine, get_db, Base, SessionLocal
from models import User, Mahsulot, Mashina, Hujjat, Olchov, HujjatHolati, HujjatRaqamHisoblagich, TizimXatosi, TahrirTarixi, Firma, MahsulotNarxi, Sozlama
from schemas import UserLogin, Token, MashinaCreate, HujjatCreate, HujjatUpdate, OlchovCreate, UserCreate, UserParolYangilash, UserHolatYangilash, FirmaCreate, MahsulotNarxiCreate, PinOrnatish, PinTekshirish, TaroziYubor
from auth import verify_password, create_access_token, hash_password, get_current_user, require_role, require_moliyaviy_ruxsat
from config import PG_DUMP_YOL, KAMERA_1_IP, KAMERA_2_IP, KAMERA_LOGIN, KAMERA_PAROL, SERVER_ASOSIY_URL, ALLOWED_ORIGINS, DATABASE_URL, TARMOQ_BACKUP_IP, TARMOQ_BACKUP_SHARE, TARMOQ_BACKUP_FOYDALANUVCHI, TARMOQ_BACKUP_PAROL, TAROZI_AGENT_KEY
from utils import konditsion_hisobla, xavfsiz_papka_nomi, xavfsiz_sana
from datetime import datetime
import html
import io
import threading
import time
from pathlib import Path
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.errors import RateLimitExceeded
from slowapi.middleware import SlowAPIMiddleware

Base.metadata.create_all(bind=engine)

app = FastAPI(title="Hazorasp Tekstil Tarozi Tizimi", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    # Ilova cookie emas, Authorization: Bearer sarlavhasidan foydalanadi -
    # shuning uchun credentials=True kerak emas (bu faqat cookie-asosidagi
    # oqimlarga tegishli), False qilib qo'yish keraksiz xavfni yopadi.
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
    # Chrome "Private Network Access": Cloudflare Tunnel orqasidagi origin
    # ba'zan mijoz brauzeriga "private" tarmoq sifatida ko'rinadi, shunda
    # Chrome preflight'ga Access-Control-Request-Private-Network: true
    # qo'shadi - buni yoqmasak, Starlette CORSMiddleware 400 "Disallowed
    # CORS private-network" bilan preflight'ni rad etadi (masalan telefon
    # ilovadan mobil tarmoq orqali kirganda POST /login shu tarzda buzilgan).
    allow_private_network=True,
)


def _mijoz_ip(request: Request) -> str:
    """IP-boshiga tezlik cheklovi uchun haqiqiy mijoz IP'sini aniqlaydi.
    XAVFSIZLIK: `X-Forwarded-For` ATAYLAB ishlatilmaydi - bu sarlavha
    mijozning o'zi tomonidan erkin, cheklovsiz sozlanadi (oddiy HTTP
    so'rov sarlavhasi), shuning uchun har bir so'rovda boshqa-boshqa
    soxta qiymat yuborib, tezlik-cheklovini (login/PIN brute-force
    himoyasi) BUTUNLAY chetlab o'tish mumkin edi - bu real sinovda
    tasdiqlangan haqiqiy zaiflik edi.
    Backend doimo Cloudflare Tunnel orqasida ishlaydi - Cloudflare
    edge'ining o'zi qo'yadigan `CF-Connecting-IP` sarlavhasini mijoz
    SOXTALASHTIRA OLMAYDI (Cloudflare buni har doim o'zi, haqiqiy
    ulanish asosida qayta yozadi). Agar bu sarlavha yo'q bo'lsa
    (masalan lokal tarmoqdan to'g'ridan-to'g'ri ulanish), oddiy TCP
    ulanish IP'siga qaytiladi."""
    cf_ip = request.headers.get("cf-connecting-ip")
    if cf_ip:
        return cf_ip.strip()
    return request.client.host if request.client else "noma-lum"


limiter = Limiter(key_func=_mijoz_ip)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)
app.add_middleware(SlowAPIMiddleware)

# ============ TAROZI (alohida "Tarozi agenti" dasturidan tarmoq orqali qabul qilinadi) ============
# Tarozi (KLAS XK3190-A9+) tarozixonadagi ALOHIDA kompyuterga ulangan -
# server esa ofisda turadi. Shuning uchun server COM port bilan bevosita
# ishlamaydi: tarozixonadagi kompyuterda ishlaydigan tarozi_agent.py COM
# portni o'qib, har necha yuz millisekundda quyidagi endpointga (POST
# /tarozi/yubor) so'nggi qiymatni yuboradi. Server faqat shu qiymatni
# xotirada saqlab, GET /tarozi/joriy orqali frontendga beradi.
_TAROZI_ULANISH_TIMEOUT_SONIYA = 10.0

_tarozi_qulf = threading.Lock()
_tarozi_holat = {"ogirlik_kg": 0.0, "ulangan": False}
_tarozi_oxirgi_yangilanish = 0.0


@app.post("/tarozi/yubor")
def tarozi_qiymat_qabul_qil(
    payload: TaroziYubor,
    x_tarozi_agent_key: str = Header(default=""),
):
    """Tarozi agentidan (tarozixonadagi kompyuter) qabul qiladi. Foydalanuvchi
    JWT'i emas - alohida umumiy sir (TAROZI_AGENT_KEY) bilan himoyalangan,
    chunki bu kanalda login qiluvchi odam emas, bitta ishonchli qurilma bor."""
    if not TAROZI_AGENT_KEY or x_tarozi_agent_key != TAROZI_AGENT_KEY:
        raise HTTPException(status_code=401, detail="Noto'g'ri agent kaliti")

    global _tarozi_oxirgi_yangilanish
    with _tarozi_qulf:
        _tarozi_holat["ogirlik_kg"] = payload.ogirlik_kg
        _tarozi_holat["ulangan"] = payload.ulangan
        _tarozi_oxirgi_yangilanish = time.time()
    return {"ok": True}


@app.get("/tarozi/joriy")
def tarozi_joriy_ogirlik(current_user: dict = Depends(get_current_user)):
    with _tarozi_qulf:
        holat = dict(_tarozi_holat)
        yangi_malumotmi = (time.time() - _tarozi_oxirgi_yangilanish) <= _TAROZI_ULANISH_TIMEOUT_SONIYA
    # Agentdan 10 soniyadan ko'proq vaqt hech narsa kelmagan bo'lsa - agent
    # yoki tarmoq o'zi o'lgan deb hisoblanadi, hatto agentning oxirgi
    # yuborgan holati "ulangan: true" bo'lsa ham.
    if not yangi_malumotmi:
        holat["ulangan"] = False
    return holat

# ============ ASOSIY ============

@app.get("/")
def root():
    return {"message": "Hazorasp Tekstil Tarozi Tizimi ishlamoqda!"}

@app.get("/health")
def health():
    return {"status": "ok"}

@app.post("/login", response_model=Token)
@limiter.limit("5/minute")
def login(request: Request, user_data: UserLogin, db: Session = Depends(get_db)):
    user = db.query(User).filter(
        User.username == user_data.username,
        User.role == user_data.role,
        User.is_active == True,  # noqa: E712 - SQLAlchemy filtrida `is` emas `==` kerak
    ).first()
    # DIQQAT: o'chirilgan (is_active=False) hisob ham AYNAN shu umumiy
    # "Login yoki parol noto'g'ri" xabarini olishi kerak - alohida xabar
    # (masalan "hisob o'chirilgan") hisob mavjudligini oshkor qilib
    # qo'yadi (username enumeration).
    if not user or not verify_password(user_data.password, user.password):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Login yoki parol noto'g'ri!"
        )
    token = create_access_token({"sub": user.username, "role": user.role, "id": user.id})
    return {"access_token": token, "token_type": "bearer", "role": user.role, "username": user.username}

@app.post("/setup")
def setup(db: Session = Depends(get_db)):
    # Autentifikatsiyasiz endpoint - lekin bu muammo emas, chunki yangi
    # (bo'sh) bazada hali hech qanday token/parol yo'q, autentifikatsiya
    # qo'shib bo'lmaydi (tuxum-tovuq muammosi). Shu sabab himoya IDENTITY
    # orqali emas, BAZA HOLATI orqali qilinadi: agar bazada ALLAQACHON
    # kamida bitta foydalanuvchi (istalgan rolda) bo'lsa, bu - haqiqiy,
    # ishlab turgan tizim, va standart (hammaga ma'lum) admin/operator
    # parollarini qayta yoki qo'shimcha yaratishga urinish rad etiladi.
    bironta_user = db.query(User).first()
    if bironta_user:
        raise HTTPException(
            status_code=403,
            detail="Tizim allaqachon sozlangan - /setup faqat bo'sh bazada, "
                   "birinchi marta ishga tushirishda ishlaydi.",
        )
    new_admin = User(
        username="admin",
        password=hash_password("admin123"),
        role="admin",
        is_active=True
    )
    db.add(new_admin)
    new_operator = User(
        username="operator",
        password=hash_password("operator123"),
        role="operator",
        is_active=True
    )
    db.add(new_operator)
    mahsulotlar = [
        Mahsulot(nom="Chigit", konditsiya_bor=True),
        Mahsulot(nom="Chiganoq", konditsiya_bor=False),
        Mahsulot(nom="Chiganoq po'chog'i", konditsiya_bor=False),
    ]
    for m in mahsulotlar:
        db.add(m)
    db.commit()
    return {"message": "Tizim sozlandi!"}

# ============ FOYDALANUVCHILAR ============

@app.get("/users")
def users_royxati(db: Session = Depends(get_db), current_user: dict = Depends(require_role("admin"))):
    foydalanuvchilar = db.query(User).order_by(User.id).all()
    return [
        {
            "id": f.id,
            "username": f.username,
            "role": f.role,
            "is_active": f.is_active,
        }
        for f in foydalanuvchilar
    ]

@app.post("/users")
def user_qoshish(user: UserCreate, db: Session = Depends(get_db), current_user: dict = Depends(require_role("admin"))):
    mavjud = db.query(User).filter(User.username == user.username).first()
    if mavjud:
        raise HTTPException(status_code=409, detail=f"'{user.username}' logini allaqachon band!")
    yangi = User(
        username=user.username,
        password=hash_password(user.password),
        role=user.role,
        is_active=True,
    )
    db.add(yangi)
    db.commit()
    db.refresh(yangi)
    return {"id": yangi.id, "username": yangi.username, "role": yangi.role, "is_active": yangi.is_active}

@app.put("/users/{user_id}/parol")
def user_parolini_yangilash(user_id: int, data: UserParolYangilash, db: Session = Depends(get_db), current_user: dict = Depends(require_role("admin"))):
    foydalanuvchi = db.query(User).filter(User.id == user_id).first()
    if not foydalanuvchi:
        raise HTTPException(status_code=404, detail="Foydalanuvchi topilmadi!")
    foydalanuvchi.password = hash_password(data.yangi_parol)
    db.commit()
    return {"status": "ok"}

@app.put("/users/{user_id}/holat")
def user_holatini_yangilash(user_id: int, data: UserHolatYangilash, db: Session = Depends(get_db), current_user: dict = Depends(require_role("admin"))):
    foydalanuvchi = db.query(User).filter(User.id == user_id).first()
    if not foydalanuvchi:
        raise HTTPException(status_code=404, detail="Foydalanuvchi topilmadi!")

    if not data.is_active:
        # 1-himoya: admin o'zini-o'zi faolsizlantira olmaydi - aks holda
        # tizimga umuman kira olmay qolishi mumkin.
        if foydalanuvchi.id == current_user.get("id"):
            raise HTTPException(status_code=400, detail="O'zingizni faolsizlantira olmaysiz!")

        # 2-himoya (qo'shimcha ehtiyot chorasi): bu amaldan keyin tizimda
        # bironta ham faol admin qolmasligi mumkin bo'lsa - rad etiladi.
        # (1-himoya tufayli bu holat amalda deyarli sodir bo'lmaydi, chunki
        # amalni bajarayotgan admin har doim faol qoladi, lekin kelajakda
        # kod o'zgarsa ham tizim qulflanib qolmasligi uchun qo'shimcha
        # tekshiruv sifatida saqlanadi.)
        if foydalanuvchi.role == "admin":
            faol_adminlar = db.query(User).filter(
                User.role == "admin",
                User.is_active == True,  # noqa: E712
                User.id != foydalanuvchi.id,
            ).count()
            if faol_adminlar == 0:
                raise HTTPException(status_code=400, detail="Oxirgi faol adminni faolsizlantirib bo'lmaydi!")

    foydalanuvchi.is_active = data.is_active
    db.commit()
    db.refresh(foydalanuvchi)
    return {
        "id": foydalanuvchi.id,
        "username": foydalanuvchi.username,
        "role": foydalanuvchi.role,
        "is_active": foydalanuvchi.is_active,
    }

# ============ MASHINALAR ============

@app.post("/mashinalar")
def mashina_qoshish(mashina: MashinaCreate, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    mavjud = db.query(Mashina).filter(
        Mashina.davlat_raqami == mashina.davlat_raqami
    ).first()
    if mavjud:
        return mavjud
    yangi = Mashina(**mashina.dict())
    db.add(yangi)
    firma_royxatga_qoshish(db, mashina.firma)
    db.commit()
    db.refresh(yangi)
    return yangi

@app.get("/mashinalar")
def mashinalar_royxati(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    return db.query(Mashina).all()

@app.get("/mashinalar/qidiruv/{raqam}")
def mashina_qidiruv(raqam: str, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    return db.query(Mashina).filter(
        Mashina.davlat_raqami.ilike(f"%{raqam}%")
    ).all()

# ============ FIRMALAR ============
# Faqat qo'shimcha (additive) tanlov ro'yxati - operator ekranidagi
# autocomplete uchun. Mashina.firma/Hujjat.firma ustunlariga bog'liq
# emas (FK yo'q), ular hozirgidek erkin matn bo'lib qoladi.

def firma_royxatga_qoshish(db: Session, nom: str | None):
    """Agar `nom` bo'sh bo'lmasa va Firma jadvalida (katta-kichik harf/
    probelga sezgir bo'lmagan holda) hali yo'q bo'lsa - yangi qator
    sifatida qo'shadi. Chaqiruvchi keyin o'zi commit qiladi (bu funksiya
    faqat db.add() qiladi, tranzaksiyani boshqarmaydi)."""
    if not nom or not nom.strip():
        return
    toza_nom = nom.strip()
    mavjud = db.query(Firma).filter(func.lower(Firma.nom) == toza_nom.lower()).first()
    if mavjud:
        return
    db.add(Firma(nom=toza_nom))


@app.get("/firmalar")
def firmalar_royxati(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    return db.query(Firma).order_by(Firma.nom).all()


@app.post("/firmalar")
def firma_qoshish(firma: FirmaCreate, db: Session = Depends(get_db), current_user: dict = Depends(require_role("admin"))):
    mavjud = db.query(Firma).filter(func.lower(Firma.nom) == firma.nom.lower()).first()
    if mavjud:
        raise HTTPException(status_code=409, detail=f"'{firma.nom}' firmasi allaqachon ro'yxatda bor!")
    yangi = Firma(nom=firma.nom)
    db.add(yangi)
    db.commit()
    db.refresh(yangi)
    return yangi

# ============ MAHSULOTLAR ============

@app.get("/mahsulotlar")
def mahsulotlar_royxati(db: Session = Depends(get_db)):
    # Login'dan OLDINGI mahsulot-tanlash ekrani shu endpointni tokensiz chaqiradi,
    # shuning uchun bu yerda autentifikatsiya talab qilinmaydi.
    return db.query(Mahsulot).filter(Mahsulot.is_active == True).all()

# Narx - moliyaviy ma'lumot, shuning uchun oddiy "admin" roli YETARLI
# EMAS - require_moliyaviy_ruxsat() orqali alohida PIN-tokeni talab
# qilinadi (qarang: auth.py, POST /moliyaviy/pin-tekshir).
@app.post("/mahsulotlar/{mahsulot_id}/narx")
def mahsulot_narxi_qoshish(mahsulot_id: int, data: MahsulotNarxiCreate, db: Session = Depends(get_db), current_user: dict = Depends(require_moliyaviy_ruxsat)):
    mahsulot = db.query(Mahsulot).filter(Mahsulot.id == mahsulot_id).first()
    if not mahsulot:
        raise HTTPException(status_code=404, detail="Mahsulot topilmadi!")
    yangi = MahsulotNarxi(mahsulot_id=mahsulot_id, narx=data.narx)
    db.add(yangi)
    db.commit()
    db.refresh(yangi)
    return yangi

@app.get("/mahsulotlar/narxlar")
def mahsulotlar_narxlari(db: Session = Depends(get_db), current_user: dict = Depends(require_moliyaviy_ruxsat)):
    mahsulotlar = db.query(Mahsulot).filter(Mahsulot.is_active == True).all()
    natija = []
    for m in mahsulotlar:
        tarix = db.query(MahsulotNarxi).filter(
            MahsulotNarxi.mahsulot_id == m.id
        ).order_by(MahsulotNarxi.created_at.desc()).all()
        natija.append({
            "mahsulot_id": m.id,
            "mahsulot_nomi": m.nom,
            "hozirgi_narx": tarix[0].narx if tarix else None,
            "tarix": [{"id": t.id, "narx": t.narx, "created_at": str(t.created_at)} for t in tarix],
        })
    return natija

# ============ MOLIYAVIY HISOBOT ============
# PIN bilan himoyalangan bo'lim - qarang: auth.py require_moliyaviy_ruxsat().
# PIN xeshi Sozlama jadvalida "moliyaviy_pin_hash" kaliti ostida
# saqlanadi (umumiy GET /sozlamalar javobidan ATAYLAB chiqarib
# tashlanadi - qarang pastda sozlamalar_olish()).

MOLIYAVIY_PIN_SOZLAMA_KALIT = "moliyaviy_pin_hash"
MOLIYAVIY_TOKEN_MUDDATI_DAQIQA = 20

@app.put("/moliyaviy/pin")
@limiter.limit("5/minute")
def moliyaviy_pin_ornatish(request: Request, data: PinOrnatish, db: Session = Depends(get_db), current_user: dict = Depends(require_role("admin"))):
    sozlama = db.query(Sozlama).filter(Sozlama.kalit == MOLIYAVIY_PIN_SOZLAMA_KALIT).first()
    if sozlama and sozlama.qiymat:
        # PIN allaqachon o'rnatilgan - ESKI PIN to'g'ri kiritilishi
        # SHART, aks holda PIN'ni bilmagan admin ham uni o'ziga mos
        # qilib qayta o'rnatib, butun himoyani chetlab o'tishi mumkin edi.
        if not data.eski_pin or not verify_password(data.eski_pin, sozlama.qiymat):
            raise HTTPException(status_code=401, detail="Eski PIN noto'g'ri!")
        sozlama.qiymat = hash_password(data.yangi_pin)
        sozlama.updated_at = datetime.now()
    else:
        # Birinchi marta o'rnatish - eski PIN talab qilinmaydi.
        db.add(Sozlama(kalit=MOLIYAVIY_PIN_SOZLAMA_KALIT, qiymat=hash_password(data.yangi_pin)))
    db.commit()
    return {"status": "ok"}

@app.post("/moliyaviy/pin-tekshir")
@limiter.limit("5/minute")
def moliyaviy_pin_tekshir(request: Request, data: PinTekshirish, db: Session = Depends(get_db), current_user: dict = Depends(require_role("admin"))):
    sozlama = db.query(Sozlama).filter(Sozlama.kalit == MOLIYAVIY_PIN_SOZLAMA_KALIT).first()
    if not sozlama or not sozlama.qiymat:
        raise HTTPException(status_code=400, detail="PIN hali o'rnatilmagan - avval PIN o'rnating!")
    if not verify_password(data.pin, sozlama.qiymat):
        raise HTTPException(status_code=401, detail="PIN noto'g'ri!")
    token = create_access_token(
        {"sub": current_user["sub"], "role": current_user["role"], "id": current_user["id"],
         "moliyaviy_ruxsat": True},
        expires_minutes=MOLIYAVIY_TOKEN_MUDDATI_DAQIQA,
    )
    return {"moliyaviy_token": token, "muddat_daqiqa": MOLIYAVIY_TOKEN_MUDDATI_DAQIQA}


def _moliyaviy_davr_boshi(davr: str):
    from datetime import date, timedelta
    bugun = date.today()
    if davr == "kunlik":
        return bugun
    if davr == "haftalik":
        return bugun - timedelta(days=7)
    if davr == "oylik":
        return bugun.replace(day=1)
    if davr == "mavsum":
        # Mavsum: 1 Avgust dan 31 Iyul gacha - /statistika/mavsum bilan
        # BIR XIL qoida (Excel jurnaldagi kalendar-yil qoidasidan farqli).
        if bugun.month >= 8:
            return date(bugun.year, 8, 1)
        return date(bugun.year - 1, 8, 1)
    raise HTTPException(status_code=400, detail="Davr faqat kunlik/haftalik/oylik/mavsum bo'lishi mumkin!")


@app.get("/moliyaviy/hisobot")
def moliyaviy_hisobot(davr: str, db: Session = Depends(get_db), current_user: dict = Depends(require_moliyaviy_ruxsat)):
    davr_boshi = _moliyaviy_davr_boshi(davr)

    # Har hujjat uchun alohida netto/konditsion jami - N+1 so'rovsiz,
    # bitta JOIN+GROUP BY orqali (Olchov.hujjat_id bo'yicha).
    satrlar = db.query(
        Hujjat.id, Hujjat.mahsulot_id, Hujjat.created_at,
        func.coalesce(func.sum(Olchov.netto), 0).label('netto'),
        func.coalesce(func.sum(Olchov.konditsion), 0).label('konditsion'),
    ).outerjoin(Olchov, Olchov.hujjat_id == Hujjat.id).filter(
        Hujjat.holat == HujjatHolati.TUGALLANDI,
        Hujjat.created_at >= davr_boshi,
    ).group_by(Hujjat.id, Hujjat.mahsulot_id, Hujjat.created_at).all()

    mahsulotlar = {m.id: m for m in db.query(Mahsulot).all()}

    # Barcha narx yozuvlarini BIR MARTA olib, mahsulot bo'yicha
    # (sana o'sish tartibida) xotirada guruhlash - har hujjat uchun
    # alohida DB so'rovi shart emas.
    barcha_narxlar = db.query(MahsulotNarxi).order_by(
        MahsulotNarxi.mahsulot_id, MahsulotNarxi.created_at).all()
    narxlar_mahsulot_boyicha = {}
    for n in barcha_narxlar:
        narxlar_mahsulot_boyicha.setdefault(n.mahsulot_id, []).append(n)

    def _narxni_top(mahsulot_id, sana):
        mos_narx = None
        for n in narxlar_mahsulot_boyicha.get(mahsulot_id, []):
            if n.created_at <= sana:
                mos_narx = n.narx
            else:
                break
        return mos_narx

    natija = {}
    for satr in satrlar:
        mahsulot = mahsulotlar.get(satr.mahsulot_id)
        if not mahsulot:
            continue
        asos_kg = satr.konditsion if mahsulot.konditsiya_bor else satr.netto

        yozuv = natija.setdefault(satr.mahsulot_id, {
            "mahsulot_id": satr.mahsulot_id,
            "mahsulot_nomi": mahsulot.nom,
            "asos": "konditsion" if mahsulot.konditsiya_bor else "netto",
            "jami_kg": 0.0,
            "daromad": 0.0,
            "narxsiz_hujjatlar_soni": 0,
        })
        yozuv["jami_kg"] += asos_kg

        narx = _narxni_top(satr.mahsulot_id, satr.created_at)
        if narx is None:
            yozuv["narxsiz_hujjatlar_soni"] += 1
            continue
        yozuv["daromad"] += asos_kg * narx

    for yozuv in natija.values():
        yozuv["jami_kg"] = round(yozuv["jami_kg"], 2)
        yozuv["daromad"] = round(yozuv["daromad"], 2)

    return {
        "davr": davr,
        "davr_boshi": str(davr_boshi),
        "mahsulotlar": list(natija.values()),
        "jami_daromad": round(sum(v["daromad"] for v in natija.values()), 2),
    }

# ============ HUJJATLAR ============

MAHSULOT_RAQAM_PREFIKS = {1: "CHG", 2: "CHN", 3: "CHP", 4: "PTZ"}

def keyingi_hujjat_raqami(db: Session, yil: int, mahsulot_id: int) -> str:
    hisoblagich = db.query(HujjatRaqamHisoblagich).filter(
        HujjatRaqamHisoblagich.yil == yil,
        HujjatRaqamHisoblagich.mahsulot_id == mahsulot_id
    ).with_for_update().first()

    if not hisoblagich:
        hisoblagich = HujjatRaqamHisoblagich(yil=yil, mahsulot_id=mahsulot_id, oxirgi_raqam=0)
        db.add(hisoblagich)
        db.flush()
        hisoblagich = db.query(HujjatRaqamHisoblagich).filter(
            HujjatRaqamHisoblagich.yil == yil,
            HujjatRaqamHisoblagich.mahsulot_id == mahsulot_id
        ).with_for_update().first()

    hisoblagich.oxirgi_raqam += 1
    db.flush()
    prefiks = MAHSULOT_RAQAM_PREFIKS.get(mahsulot_id, "DOC")
    return f"{prefiks}-{yil}/{str(hisoblagich.oxirgi_raqam).zfill(3)}"

@app.post("/hujjatlar")
def hujjat_yaratish(hujjat: HujjatCreate, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    if hujjat.mijoz_kaliti:
        mavjud = db.query(Hujjat).filter(Hujjat.mijoz_kaliti == hujjat.mijoz_kaliti).first()
        if mavjud:
            # Bu so'rov allaqachon bajarilgan (masalan javob yo'qolib,
            # frontend qayta yuborgan) - yangisini yaratmasdan mavjudini
            # qaytaramiz, shunda ikkilamchi hujjat/raqam sarflanmaydi.
            return mavjud

    # mahsulot_id/mashina_id noto'g'ri (mavjud bo'lmagan) bo'lsa, bazaning
    # FK cheklovi keyinroq xom holda otilib, chiroyli xabar o'rniga 500
    # xato berardi - shu sabab bu yerda OLDINDAN, hech qanday yon ta'sir
    # (masalan hujjat raqami hisoblagichini oshirish) sodir bo'lishidan
    # avval tekshiriladi.
    mahsulot = db.query(Mahsulot).filter(Mahsulot.id == hujjat.mahsulot_id).first()
    if not mahsulot:
        raise HTTPException(
            status_code=404, detail=f"Mahsulot topilmadi (id={hujjat.mahsulot_id})")
    mashina = db.query(Mashina).filter(Mashina.id == hujjat.mashina_id).first()
    if not mashina:
        raise HTTPException(
            status_code=404, detail=f"Mashina topilmadi (id={hujjat.mashina_id})")

    yil = datetime.now().year
    yangi_raqam = keyingi_hujjat_raqami(db, yil, hujjat.mahsulot_id)
    yangi = Hujjat(
        raqam=yangi_raqam,
        mashina_raqami=mashina.davlat_raqami,
        shofyor=mashina.shofyor,
        firma=mashina.firma,
        **hujjat.dict(),
    )
    db.add(yangi)
    db.commit()
    db.refresh(yangi)
    return yangi

def _olchovlar_jamlangan(olchovlar):
    """Hujjatning barcha Olchov qatorlaridan JAMI tara/brutto/netto/konditsion
    hisoblaydi. Bitta arava uchun bir nechta qator bo'lishi normal holat
    (operator avval faqat TARA, keyin TARA+BRUTTO birga saqlaydi - shu sabab
    shu arava_raqam bo'yicha ikkinchi qator birinchisini "davom ettiradi",
    ikkalasi ham alohida arava sifatida qo'shilmasligi kerak). Shu sabab
    avval har arava_raqam uchun bitta "jamlangan" qiymat (har maydon bo'yicha
    eng oxirgi NULL bo'lmagan qiymat) hisoblanadi, SO'NGRA shu jamlangan
    aravalar bo'yicha yig'indi olinadi - xuddi nakladnoy_uchun_malumot()dagi
    bilan bir xil mantiq (bu yerda takrorlanish tufayli ikki baravar
    hisoblanmasligi uchun).
    """
    aravalar = {}
    for o in sorted(olchovlar, key=lambda x: x.id):
        a = aravalar.setdefault(o.arava_raqam, {
            "tara": None, "brutto": None, "netto": None, "konditsion": None,
        })
        for maydon in ("tara", "brutto", "netto", "konditsion"):
            qiymat = getattr(o, maydon)
            if qiymat is not None:
                a[maydon] = qiymat

    jami_tara = sum(a["tara"] for a in aravalar.values() if a["tara"]) or None
    jami_brutto = sum(a["brutto"] for a in aravalar.values() if a["brutto"]) or None
    jami_netto = sum(a["netto"] for a in aravalar.values() if a["netto"]) or None
    jami_konditsion = sum(a["konditsion"] for a in aravalar.values() if a["konditsion"]) or None
    return jami_tara, jami_brutto, jami_netto, jami_konditsion


def _hujjat_navbat_fallback(hujjat_qiymat, navbat, maydon_nomi):
    """Operator ekrani klass/sinf/seleksiya_navi/terim_turi/tiket_raqam/
    tuda_raqam ni to'g'ridan-to'g'ri Hujjat'ga emas, alohida Navbat
    jadvaliga saqlaydi (qarang: nakladnoy_uchun_malumot() boshidagi
    izoh) - Hujjat'ning o'z qiymati bo'sh bo'lsa, Navbat'dagi mos
    maydonga fallback qilinadi. Excel jurnal va Nakladnoy PDF bu
    fallback'ni allaqachon qo'llaydi - bu yerda xuddi shu naqsh
    GET /hujjatlar va GET /hujjatlar/{id}ga (demak admin "Tuzat"
    oynasiga) ham qo'llaniladi."""
    if not _boshmi(hujjat_qiymat):
        return hujjat_qiymat
    if navbat is not None:
        navbat_qiymat = getattr(navbat, maydon_nomi)
        if not _boshmi(navbat_qiymat):
            return navbat_qiymat
    return None


@app.get("/hujjatlar")
def hujjatlar_royxati(
    mahsulot_id: int = None,
    bekor_qilinganlarni_korsat: bool = False,
    sana_dan: str = None,
    sana_gacha: str = None,
    sahifa: int = 1,
    sahifa_hajmi: int = 50,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    so_rov = db.query(Hujjat)
    if mahsulot_id:
        so_rov = so_rov.filter(Hujjat.mahsulot_id == mahsulot_id)
    if not bekor_qilinganlarni_korsat:
        so_rov = so_rov.filter(Hujjat.holat != HujjatHolati.BEKOR_QILINDI)
    if sana_dan:
        so_rov = so_rov.filter(Hujjat.created_at >= sana_dan)
    if sana_gacha:
        so_rov = so_rov.filter(Hujjat.created_at < sana_gacha)

    jami_soni = so_rov.count()

    hujjatlar = so_rov.order_by(Hujjat.id.desc()) \
        .offset((sahifa - 1) * sahifa_hajmi) \
        .limit(sahifa_hajmi) \
        .all()

    # Operator ekrani klass/sinf/seleksiya_navi/terim_turi/tiket_raqam/
    # tuda_raqam ni Hujjat'ga emas, Navbat'ga saqlagani uchun (qarang:
    # _hujjat_navbat_fallback()) - shu sahifadagi barcha hujjat uchun
    # Navbat qatorlarini BIR so'rovda (N+1 qilmasdan) oldindan olamiz.
    from models import Navbat
    hujjat_idlar = [h.id for h in hujjatlar]
    navbat_dict = {
        n.hujjat_id: n
        for n in db.query(Navbat).filter(Navbat.hujjat_id.in_(hujjat_idlar)).all()
    }

    natija = []
    for h in hujjatlar:
        olchovlar = db.query(Olchov).filter(Olchov.hujjat_id == h.id).order_by(Olchov.id.asc()).all()
        jami_tara, jami_brutto, jami_netto, jami_konditsion = _olchovlar_jamlangan(olchovlar)
        # namlik/ifloslik Hujjatda emas, Olchov qatorlarida saqlanadi -
        # eksport uchun eng oxirgi NULL bo'lmagan qiymat olinadi
        # (nakladnoy_uchun_malumot()dagi bilan bir xil mantiq).
        namlik = None
        ifloslik = None
        for o in olchovlar:
            if o.namlik is not None:
                namlik = o.namlik
            if o.ifloslik is not None:
                ifloslik = o.ifloslik
        navbat = navbat_dict.get(h.id)
        natija.append({
            "id": h.id,
            "raqam": h.raqam,
            "mashina_id": h.mashina_id,
            "mashina_raqami": h.mashina_raqami or "—",
            "shofyor": h.shofyor or "—",
            "firma": h.firma or "—",
            "mahsulot_id": h.mahsulot_id,
            "aravalar_soni": h.aravalar_soni,
            "tuda_raqam": _hujjat_navbat_fallback(h.tuda_raqam, navbat, "tuda_raqam"),
            "tiket_raqam": _hujjat_navbat_fallback(h.tiket_raqam, navbat, "tiket_raqam"),
            "klass": _hujjat_navbat_fallback(h.klass, navbat, "klass"),
            "sinf": _hujjat_navbat_fallback(h.sinf, navbat, "sinf"),
            "seleksiya_navi": _hujjat_navbat_fallback(h.seleksiya_navi, navbat, "seleksiya_navi"),
            "terim_turi": _hujjat_navbat_fallback(h.terim_turi, navbat, "terim_turi"),
            "qabul_qildi": h.qabul_qildi,
            "yuk_olindi": h.yuk_olindi,
            "holat": h.holat,
            "tara": jami_tara,
            "brutto": jami_brutto,
            "netto": jami_netto,
            "konditsion": jami_konditsion,
            "namlik": namlik,
            "ifloslik": ifloslik,
            "created_at": str(h.created_at) if h.created_at else None,
        })
    return {
        "natijalar": natija,
        "jami": jami_soni,
        "sahifa": sahifa,
        "sahifa_hajmi": sahifa_hajmi,
    }


AY_NOMLARI = {
    1: "Yanvar", 2: "Fevral", 3: "Mart", 4: "Aprel", 5: "May", 6: "Iyun",
    7: "Iyul", 8: "Avgust", 9: "Sentabr", 10: "Oktabr", 11: "Noyabr", 12: "Dekabr",
}

_EKSPORT_USTUNLAR = [
    "№", "№ Naklad", "Mahsulot nomi", "Brutto (kg)", "Tara (kg)",
    "Netto (kg)", "Kondicion (kg)", "Mashina raqami", "Yuk oluvchi",
]


def _eksport_qator_yoz(ws, qiymatlar, qalin=False, fon=None):
    ws.append(qiymatlar)
    qator = ws.max_row
    for hujayra in ws[qator]:
        if qalin:
            hujayra.font = Font(bold=True)
        if fon:
            hujayra.fill = PatternFill(start_color=fon, end_color=fon, fill_type="solid")
    return qator


def _eksport_jami_qator(ws, sarlavha, netto, konditsion, konditsiya_bormi, fon,
                         ustunlar_soni, sarlavha_ustun, netto_ustun, konditsion_ustun):
    """Jami (kunlik/oylik/mavsumiy) qatorni yozadi - ustun soni va
    sarlavha/netto/konditsion QAYSI ustunda chiqishi chaqiruvchi tomonidan
    beriladi, shu sabab bu bir nechta har xil ustun-tuzilishidagi
    hisobotlar (qo'lda eksport - 9 ustun, avtomatik jurnal - 10 yoki 23
    ustun) uchun QAYTA ISHLATILADI."""
    qator = [""] * ustunlar_soni
    qator[sarlavha_ustun - 1] = sarlavha
    qator[netto_ustun - 1] = round(netto) if netto else 0
    if konditsion_ustun and konditsiya_bormi:
        qator[konditsion_ustun - 1] = round(konditsion) if konditsion else ""
    _eksport_qator_yoz(ws, qator, qalin=True, fon=fon)


# Kun/oy sarlavhalari va jami qatorlarining fon ranglari - qo'lda eksport
# va avtomatik jurnal ikkalasida ham bir xil (vizual izchillik uchun).
_KUN_FON = "D9E8D3"
_JAMI_FON = "F0D878"
_OY_FON = "E8A868"
_MAVSUM_FON = "C05050"


def _kun_oy_guruhlab_yoz(ws, itemlar, konditsiya_bormi, qator_yozuvchi,
                          ustunlar_soni, sarlavha_ustun, netto_ustun,
                          konditsion_ustun, ustunlar_royxati):
    """Sana bo'yicha kun sarlavhalari + kunlik jami, oy oxirida "OY JAMI"
    va yakunda "UMUMIY JAMI" qatorlarini yozadi - qo'lda eksport
    (hujjatlar_eksport) va avtomatik jurnal (excel_qatorga_yoz)
    ikkalasida ham AYNAN BIR XIL mantiq ishlatilishi uchun umumiy
    funksiyaga chiqarilgan.

    `itemlar` - har birida "sana" (date obyekti) kaliti bo'lgan dict'lar
    ro'yxati (tartib muhim emas - bu yerning o'zi kun bo'yicha guruhlab,
    xronologik tartiblaydi).
    `qator_yozuvchi(ws, item, kun_ichidagi_tartib_raqami)` - har bir item
    uchun asosiy ma'lumot qatori(lari)ni yozadi va shu item qo'shgan
    (netto, konditsion) yig'indisini qaytaradi - bitta itemda bir nechta
    qator bo'lishi mumkin (masalan bir nechta arava), shu sabab qaytarilgan
    qiymat "necha qator yozildi"ga emas, "netto/konditsion yig'indisi"ga
    bog'liq.
    """
    kunlar = {}
    for item in itemlar:
        kunlar.setdefault(item["sana"], []).append(item)

    oy_jami = {}
    mavsum_netto = 0
    mavsum_konditsion = 0
    joriy_oy = None

    def oy_yakunla(oy_kaliti):
        yil, oy = oy_kaliti
        j = oy_jami[oy_kaliti]
        _eksport_jami_qator(
            ws, f"{AY_NOMLARI[oy]} {yil} - OY JAMI",
            j["netto"], j["konditsion"], konditsiya_bormi, _OY_FON,
            ustunlar_soni, sarlavha_ustun, netto_ustun, konditsion_ustun)

    for sana in sorted(kunlar.keys()):
        oy_kaliti = (sana.year, sana.month)
        if joriy_oy is not None and oy_kaliti != joriy_oy:
            oy_yakunla(joriy_oy)
        joriy_oy = oy_kaliti
        oy_jami.setdefault(oy_kaliti, {"netto": 0, "konditsion": 0})

        boshlangich_qator = ws.max_row + 1
        _eksport_qator_yoz(ws, [sana.strftime("%Y-%m-%d")], qalin=True, fon=_KUN_FON)
        ws.merge_cells(start_row=boshlangich_qator, start_column=1,
                        end_row=boshlangich_qator, end_column=ustunlar_soni)
        _eksport_qator_yoz(ws, ustunlar_royxati, qalin=True)

        kun_netto = 0
        kun_konditsion = 0
        for i, item in enumerate(kunlar[sana], start=1):
            netto_qoshimcha, konditsion_qoshimcha = qator_yozuvchi(ws, item, i)
            kun_netto += netto_qoshimcha
            kun_konditsion += konditsion_qoshimcha

        _eksport_jami_qator(
            ws, "Жами:", kun_netto, kun_konditsion, konditsiya_bormi, _JAMI_FON,
            ustunlar_soni, sarlavha_ustun, netto_ustun, konditsion_ustun)

        oy_jami[oy_kaliti]["netto"] += kun_netto
        oy_jami[oy_kaliti]["konditsion"] += kun_konditsion
        mavsum_netto += kun_netto
        mavsum_konditsion += kun_konditsion

    if joriy_oy is not None:
        oy_yakunla(joriy_oy)
        _eksport_jami_qator(
            ws, "UMUMIY JAMI", mavsum_netto, mavsum_konditsion, konditsiya_bormi, _MAVSUM_FON,
            ustunlar_soni, sarlavha_ustun, netto_ustun, konditsion_ustun)


@app.get("/hujjatlar/eksport")
def hujjatlar_eksport(
    mahsulot_id: int,
    sana_dan: str = None,
    sana_gacha: str = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Bitta mahsulot uchun, sana bo'yicha kun/oy guruhlangan Excel hisoboti -
    RASMLAR papkasiga saqlanadi VA baytlar sifatida to'g'ridan-to'g'ri
    qaytariladi (frontend shu baytlardan brauzerda yuklab olishni
    boshlaydi). Bekor qilingan hujjatlar chiqarib tashlanadi (GET
    /hujjatlar bilan bir xil standart xatti-harakat).

    DIQQAT: bu route /hujjatlar/{hujjat_id}dan OLDIN ro'yxatga olinishi
    SHART - FastAPI marshrutlarni ro'yxatga olingan tartibda tekshiradi,
    aks holda "/hujjatlar/eksport" so'rovi {hujjat_id}="eksport" sifatida
    quyidagi wildcard route tomonidan tutib qolinadi (xato: 422 int_parsing)."""
    mahsulot = db.query(Mahsulot).filter(Mahsulot.id == mahsulot_id).first()
    if not mahsulot:
        raise HTTPException(status_code=404, detail=f"Mahsulot topilmadi (id={mahsulot_id})")

    so_rov = db.query(Hujjat).filter(
        Hujjat.mahsulot_id == mahsulot_id,
        Hujjat.holat != HujjatHolati.BEKOR_QILINDI,
    )
    if sana_dan:
        so_rov = so_rov.filter(Hujjat.created_at >= sana_dan)
    if sana_gacha:
        so_rov = so_rov.filter(Hujjat.created_at < sana_gacha)
    hujjatlar = so_rov.order_by(Hujjat.created_at.asc()).all()

    satrlar = []
    for h in hujjatlar:
        olchovlar = db.query(Olchov).filter(Olchov.hujjat_id == h.id).order_by(Olchov.id.asc()).all()
        tara, brutto, netto, konditsion = _olchovlar_jamlangan(olchovlar)
        satrlar.append({
            "sana": h.created_at.date(),
            "raqam": h.raqam,
            "brutto": brutto, "tara": tara, "netto": netto, "konditsion": konditsion,
            "mashina_raqami": h.mashina_raqami or "—",
            "yuk_oluvchi": h.yuk_oluvchi or "—",
        })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = mahsulot.nom[:31]

    def _eksport_qator_yozuvchi(ws, s, i):
        # DIQQAT: "Жами" (va undan yuqori OY/UMUMIY jami) qatorlari
        # xom (kasr) qiymatlar emas, balki AYNAN shu qatorda ko'rsatilgan
        # (yaxlitlangan butun) qiymatlar yig'indisidan hisoblanadi - aks
        # holda "har bir qatorni qo'lda qo'shsam, Жамидан boshqacha
        # chiqadi" degan (matematik jihatdan to'g'ri, lekin auditor
        # uchun chalkash) nomuvofiqlik paydo bo'ladi.
        # Hisoblash uchun (0 - hali yo'q qiymat, yig'indiga ta'sir
        # qilmasligi kerak) va KO'RSATISH uchun (bo'sh katak - "0"
        # emas, chunki 0 kg haqiqiy o'lchovdan farqlanishi kerak)
        # qiymatlar ATAYLAB alohida hisoblanadi.
        netto_yaxlit = round(s["netto"]) if s["netto"] is not None else 0
        konditsion_yaxlit = (
            round(s["konditsion"])
            if (mahsulot.konditsiya_bor and s["konditsion"] is not None) else 0
        )
        _eksport_qator_yoz(ws, [
            i, s["raqam"], mahsulot.nom,
            round(s["brutto"]) if s["brutto"] is not None else "",
            round(s["tara"]) if s["tara"] is not None else "",
            netto_yaxlit if s["netto"] is not None else "",
            konditsion_yaxlit
            if (mahsulot.konditsiya_bor and s["konditsion"] is not None) else "",
            s["mashina_raqami"], s["yuk_oluvchi"],
        ])
        return netto_yaxlit, (konditsion_yaxlit if mahsulot.konditsiya_bor else 0)

    _kun_oy_guruhlab_yoz(
        ws, satrlar, mahsulot.konditsiya_bor, _eksport_qator_yozuvchi,
        ustunlar_soni=len(_EKSPORT_USTUNLAR),
        sarlavha_ustun=5, netto_ustun=6, konditsion_ustun=7,
        ustunlar_royxati=_EKSPORT_USTUNLAR,
    )

    kengliklar = [6, 16, 20, 12, 12, 12, 14, 16, 22]
    for i, kenglik in enumerate(kengliklar, start=1):
        ws.column_dimensions[chr(64 + i)].width = kenglik

    # Har mahsulotning barcha fayllari (rasmlar, avtomatik jurnal va shu
    # qo'lda eksport) bitta joyda - C:/RASMLAR/{Mahsulot}/ ichida - bo'lishi
    # uchun bu ham o'sha mahsulot papkasiga saqlanadi (qarang: xuddi shu
    # papkani excel_qatorga_yoz() ham ishlatadi).
    # Papka nomi rasmlar/nakladnoy va avtomatik jurnal bilan AYNAN BIR XIL
    # (probelli, asl mahsulot.nom) bo'lishi SHART - aks holda "Chiganoq
    # po'chog'i" va "Chiganoq_po'chog'i" kabi ikkita boshqa-boshqa papka
    # hosil bo'lib, bitta mahsulotning fayllari ikkiga bo'linib qoladi.
    papka = Path(f"C:/RASMLAR/{xavfsiz_papka_nomi(mahsulot.nom)}")
    papka.mkdir(parents=True, exist_ok=True)
    davr = f"{sana_dan or 'boshidan'}_{sana_gacha or 'hozirgacha'}"
    fayl_nomi = f"{mahsulot.nom.replace(' ', '_')}_{davr}.xlsx"
    wb.save(papka / fayl_nomi)

    buffer = io.BytesIO()
    wb.save(buffer)
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{fayl_nomi}"',
            # Frontend 4 ta faylni ham yuklab bo'lgach, jami nechta hujjat
            # topilganini shu sarlavhalar orqali bila oladi - shunda bo'sh
            # (hech qanday hujjat topilmagan) natijani "muvaffaqiyatli
            # yuklandi" deb noto'g'ri ko'rsatib qo'ymaydi.
            "X-Hujjatlar-Soni": str(len(satrlar)),
            "Access-Control-Expose-Headers": "X-Hujjatlar-Soni",
        },
    )


@app.get("/hujjatlar/{hujjat_id}")
def hujjat_detail(hujjat_id: int, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    hujjat = db.query(Hujjat).filter(Hujjat.id == hujjat_id).first()
    if not hujjat:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi!")
    olchovlar = db.query(Olchov).filter(Olchov.hujjat_id == hujjat.id).order_by(Olchov.id.asc()).all()
    jami_tara, jami_brutto, jami_netto, jami_konditsion = _olchovlar_jamlangan(olchovlar)
    # namlik/ifloslik Hujjatda emas, Olchov qatorlarida saqlanadi - eng
    # oxirgi NULL bo'lmagan qiymat olinadi (hujjatlar_royxati() bilan
    # bir xil mantiq).
    namlik = None
    ifloslik = None
    for o in olchovlar:
        if o.namlik is not None:
            namlik = o.namlik
        if o.ifloslik is not None:
            ifloslik = o.ifloslik

    from models import Navbat
    navbat = db.query(Navbat).filter(Navbat.hujjat_id == hujjat_id).first()

    natija = {c.name: getattr(hujjat, c.name) for c in Hujjat.__table__.columns}
    natija["tara"] = jami_tara
    natija["brutto"] = jami_brutto
    natija["netto"] = jami_netto
    natija["konditsion"] = jami_konditsion
    # Hujjat jadvalida namlik/ifloslik ustuni yo'q (faqat Olchov'da bor) -
    # yuqoridagi generic column-loop bu kalitlarni hosil qilmagan, shu
    # sabab alohida qo'shiladi.
    natija["namlik"] = namlik
    natija["ifloslik"] = ifloslik
    # Operator ekrani klass/sinf/seleksiya_navi/terim_turi/tiket_raqam/
    # tuda_raqam ni Hujjat'ga emas, Navbat'ga saqlaydi - shu sabab
    # Hujjat'ning o'zida bo'sh bo'lsa Navbat'dan to'ldiramiz.
    for maydon in ("tuda_raqam", "tiket_raqam", "klass", "sinf", "seleksiya_navi", "terim_turi"):
        natija[maydon] = _hujjat_navbat_fallback(natija[maydon], navbat, maydon)
    return natija

RUXSAT_ETILGAN_OTISHLAR = {
    HujjatHolati.JARAYON: {HujjatHolati.TUGALLANDI, HujjatHolati.BEKOR_QILINDI},
    HujjatHolati.TUGALLANDI: set(),
    HujjatHolati.BEKOR_QILINDI: set(),
}

def holat_otishi_ruxsatmi(eski: HujjatHolati, yangi: HujjatHolati) -> bool:
    if eski == yangi:
        return True
    return yangi in RUXSAT_ETILGAN_OTISHLAR.get(eski, set())

# hujjat_yangilash orqali tahrirlanishi mumkin bo'lgan va TahrirTarixi'ga
# yoziladigan barcha maydonlar (bekor_sabab bundan mustasno emas - u ham
# kuzatiladi, faqat maxsus talab qilinish qoidasi bekor_sabab_talab_qilinadi
# bilan alohida tekshiriladi).
AUDIT_MAYDONLAR = [
    'aravalar_soni', 'tuda_raqam', 'texnik_chiqit', 'sanoat_turi', 'klassifikatsiya',
    'davomlilik_raqam', 'davomlilik_dan', 'davomlilik_gacha', 'yuk_oluvchi', 'shartnoma',
    'mashina_raqami', 'shofyor', 'firma', 'tiket_raqam', 'klass', 'sinf',
    'seleksiya_navi', 'terim_turi', 'qabul_qildi', 'yuk_olindi',
    'dostaverka', 'dostaverka_vaqt',
    'holat', 'bekor_sabab',
]

def _qiymat_normal(qiymat):
    """None va bo'sh satrni bitta 'qiymat yo'q' holatiga tenglashtiradi,
    shunda frontend har doim to'liq forma yuborsa ham yolg'on o'zgarish
    aniqlanmaydi."""
    if qiymat is None:
        return None
    if isinstance(qiymat, str) and qiymat.strip() == "":
        return None
    return qiymat

def _qiymat_matn(qiymat):
    if qiymat is None:
        return None
    return qiymat.value if hasattr(qiymat, "value") else str(qiymat)


# Operator ekrani (tortish yakunlanganda) shu maydonlarni saqlaydi -
# boshqa barcha maydon (mashina_raqami, namlik, holat, va h.k.) faqat
# admin/hisobchi tomonidan, "Tuzat" oynasi orqali, sabab ko'rsatib
# o'zgartirilishi kerak.
OPERATOR_RUXSAT_ETILGAN_MAYDONLAR = {
    "qabul_qildi", "yuk_olindi", "dostaverka", "dostaverka_vaqt", "sabab",
}

@app.put("/hujjatlar/{hujjat_id}")
def hujjat_yangilash(hujjat_id: int, data: HujjatUpdate, background_tasks: BackgroundTasks, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in ("admin", "hisobchi"):
        yuborilgan = set(data.dict(exclude_unset=True).keys())
        ruxsatsiz = yuborilgan - OPERATOR_RUXSAT_ETILGAN_MAYDONLAR
        if ruxsatsiz:
            raise HTTPException(
                status_code=403,
                detail=f"Bu maydonlarni faqat admin o'zgartira oladi: {', '.join(sorted(ruxsatsiz))}"
            )
    hujjat = db.query(Hujjat).filter(Hujjat.id == hujjat_id).first()
    if not hujjat:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi!")
    eski_holat = hujjat.holat
    if data.holat is not None and data.holat != hujjat.holat:
        if not holat_otishi_ruxsatmi(hujjat.holat, data.holat):
            raise HTTPException(
                status_code=400,
                detail=f"'{hujjat.holat.value}' holatidan '{data.holat.value}' holatiga o'tish mumkin emas!"
            )

    sabab_umumiy = (data.sabab or "").strip() or None
    sabab_bekor = (data.bekor_sabab or "").strip() or None

    if data.holat == HujjatHolati.BEKOR_QILINDI:
        if not sabab_bekor:
            raise HTTPException(
                status_code=400,
                detail="Hujjatni bekor qilish uchun sabab ko'rsatilishi shart!"
            )

    payload = data.dict(exclude_unset=True)

    # namlik/ifloslik Hujjatda emas, shu hujjatga tegishli barcha Olchov
    # qatorlarida - shuning uchun AUDIT_MAYDONLAR umumiy tsiklidan tashqarida,
    # alohida ishlanadi.
    olchovlar = None
    if "namlik" in payload or "ifloslik" in payload:
        olchovlar = db.query(Olchov).filter(Olchov.hujjat_id == hujjat_id).all()
        if not olchovlar:
            raise HTTPException(
                status_code=400,
                detail="Bu hujjatda hali birorta o'lchov (arava) yo'q, namlik/ifloslik saqlab bo'lmaydi!"
            )

    ozgarishlar = []
    for maydon in AUDIT_MAYDONLAR:
        if maydon not in payload:
            continue
        eski = getattr(hujjat, maydon)
        yangi = payload[maydon]
        if _qiymat_normal(eski) == _qiymat_normal(yangi):
            continue
        ozgarishlar.append((maydon, _qiymat_matn(eski), _qiymat_matn(yangi)))

    olchov_ozgargan_maydonlar = []
    for maydon in ("namlik", "ifloslik"):
        if maydon not in payload:
            continue
        yangi_qiymat = payload[maydon]
        agar_ozgargan = any(
            _qiymat_normal(getattr(o, maydon)) != _qiymat_normal(yangi_qiymat)
            for o in olchovlar
        )
        if not agar_ozgargan:
            continue
        eski_matn = ", ".join(_qiymat_matn(getattr(o, maydon)) or "—" for o in olchovlar)
        ozgarishlar.append((maydon, eski_matn, _qiymat_matn(yangi_qiymat)))
        olchov_ozgargan_maydonlar.append(maydon)

    if ozgarishlar and not (sabab_umumiy or sabab_bekor):
        raise HTTPException(
            status_code=400,
            detail="O'zgartirish sababi ko'rsatilishi shart!"
        )

    for key, value in payload.items():
        if key in ("sabab", "namlik", "ifloslik"):
            continue
        setattr(hujjat, key, value)

    if "firma" in payload:
        firma_royxatga_qoshish(db, payload["firma"])

    if olchov_ozgargan_maydonlar:
        for maydon in olchov_ozgargan_maydonlar:
            yangi_qiymat = payload[maydon]
            for o in olchovlar:
                setattr(o, maydon, yangi_qiymat)
        for o in olchovlar:
            if o.netto and o.namlik and o.ifloslik:
                o.konditsion = konditsion_hisobla(o.netto, o.namlik, o.ifloslik)

    for maydon, eski_matn, yangi_matn in ozgarishlar:
        qator_sababi = (sabab_bekor or sabab_umumiy) if maydon in ("holat", "bekor_sabab") \
            else (sabab_umumiy or sabab_bekor)
        db.add(TahrirTarixi(
            hujjat_id=hujjat_id,
            maydon=maydon,
            eski_qiymat=eski_matn,
            yangi_qiymat=yangi_matn,
            sabab=qator_sababi,
            ozgartirgan_user_id=current_user.get("id"),
            ozgartirgan_username=current_user.get("sub"),
        ))

    db.commit()
    db.refresh(hujjat)

    # Excel jurnaliga aynan SHU YERDA, hujjat "tugallandi" holatiga
    # YANGI o'tganda yoziladi (avval POST /olchovlar'da, tara+brutto
    # saqlanganda yozilardi - lekin o'sha payt dostaverka/qabul_qildi/
    # yuk_olindi kabi maydonlar hali bo'sh bo'lardi, chunki ular ODATDA
    # aynan shu "tugallash" amali bilan bir vaqtda to'ldiriladi). Holat
    # TUGALLANDI - yakuniy (undan boshqa holatga o'tib bo'lmaydi), shu
    # sabab bu shart bitta hujjat uchun UMRIDA FAQAT BIR MARTA rost
    # bo'ladi - takroriy qator xavfi yo'q.
    if eski_holat != HujjatHolati.TUGALLANDI and hujjat.holat == HujjatHolati.TUGALLANDI:
        background_tasks.add_task(excel_qatorga_yoz_fon, hujjat_id)

    return hujjat

# ============ TAHRIR TARIXI ============

def _tahrir_yozuv_dict(y: TahrirTarixi, hujjat_raqam: str = None):
    natija = {
        "id": y.id,
        "hujjat_id": y.hujjat_id,
        "maydon": y.maydon,
        "eski_qiymat": y.eski_qiymat,
        "yangi_qiymat": y.yangi_qiymat,
        "sabab": y.sabab,
        "ozgartirgan_user_id": y.ozgartirgan_user_id,
        "ozgartirgan_username": y.ozgartirgan_username,
        "vaqt": str(y.created_at) if y.created_at else None,
    }
    if hujjat_raqam is not None:
        natija["hujjat_raqam"] = hujjat_raqam
    return natija

@app.get("/tahrirlar-tarixi")
def barcha_tahrirlar_tarixi(limit: int = 100, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in ("admin", "hisobchi"):
        raise HTTPException(status_code=403, detail="Bu amal uchun sizda ruxsat yo'q!")
    limit = min(max(limit, 1), 500)
    yozuvlar = db.query(TahrirTarixi, Hujjat.raqam).join(
        Hujjat, TahrirTarixi.hujjat_id == Hujjat.id
    ).order_by(TahrirTarixi.created_at.desc()).limit(limit).all()
    return [_tahrir_yozuv_dict(y, raqam) for y, raqam in yozuvlar]

@app.get("/tahrirlar-tarixi/{hujjat_id}")
def hujjat_tahrir_tarixi(hujjat_id: int, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in ("admin", "hisobchi"):
        raise HTTPException(status_code=403, detail="Bu amal uchun sizda ruxsat yo'q!")
    hujjat = db.query(Hujjat).filter(Hujjat.id == hujjat_id).first()
    if not hujjat:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi!")
    yozuvlar = db.query(TahrirTarixi).filter(
        TahrirTarixi.hujjat_id == hujjat_id
    ).order_by(TahrirTarixi.created_at.desc()).all()
    return [_tahrir_yozuv_dict(y) for y in yozuvlar]

# ============ OLCHOVLAR ============

@app.post("/olchovlar")
def olchov_saqlash(olchov: OlchovCreate, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    # Operator odatda shu (hujjat_id, arava_raqam) uchun avval faqat TARA,
    # keyin TARA+BRUTTO birga saqlaydi - shu sabab har safar yangi qator
    # QO'SHISH o'rniga, mavjud qator bo'lsa shu YANGILANADI (faqat so'rovda
    # kelgan, bo'sh bo'lmagan maydonlar bilan - eskisi yo'qotilmaydi). Bu
    # ham jami og'irlik ikki marta hisoblanib ketishining oldini oladi
    # (qarang: _olchovlar_jamlangan), ham offline-qayta-yuborishni xavfsiz
    # (idempotent) qiladi. Agar shu kombinatsiya uchun bir nechta eski qator
    # mavjud bo'lsa (hozircha tozalanmagan tarixiy takrorlanish), har doim
    # ENG OXIRGISI yangilanadi.
    mavjud = db.query(Olchov).filter(
        Olchov.hujjat_id == olchov.hujjat_id,
        Olchov.arava_raqam == olchov.arava_raqam,
    ).order_by(Olchov.id.desc()).first()

    if mavjud:
        yangi = mavjud
        maydonlar = olchov.dict()
        for maydon in ("tara", "brutto", "namlik", "ifloslik"):
            qiymat = maydonlar.get(maydon)
            if qiymat is not None:
                setattr(yangi, maydon, qiymat)
        yangi.qolda_kiritildi = olchov.qolda_kiritildi
    else:
        yangi = Olchov(**olchov.dict())
        db.add(yangi)

    if yangi.brutto and yangi.tara:
        yangi.netto = yangi.brutto - yangi.tara
        if yangi.namlik and yangi.ifloslik:
            yangi.konditsion = konditsion_hisobla(yangi.netto, yangi.namlik, yangi.ifloslik)
    db.add(yangi)
    db.commit()
    db.refresh(yangi)
    # DIQQAT: bu yerda ENDI excel_qatorga_yoz() chaqirilmaydi - dostaverka/
    # qabul_qildi/yuk_olindi kabi maydonlar hali bo'sh bo'ladi (ular
    # keyinroq, "tortish yakunlanganda" alohida PUT /hujjatlar/{id} orqali
    # to'ldiriladi). Yozish endi hujjat_yangilash()da, holat "tugallandi"ga
    # o'tgan PAYTDA, BARCHA maydon to'liq bo'lganda sodir bo'ladi.
    return yangi

@app.get("/olchovlar/{hujjat_id}")
def olchovlar_royxati(hujjat_id: int, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    return db.query(Olchov).filter(Olchov.hujjat_id == hujjat_id).all()

# ============ NAVBAT (PostgreSQL) ============
import json

@app.post("/navbat/qosh")
def navbat_qosh(data: dict, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    from models import Navbat
    # DIQQAT: eski yozuvni o'chirish va yangisini qo'shish BITTA
    # tranzaksiyada, bitta commit bilan bajariladi (avval ikkita alohida
    # commit bor edi - agar ikkinchisi muvaffaqiyatsiz bo'lsa, mashina
    # eski yozuv o'chirilgani uchun navbatdan butunlay g'oyib bo'lib
    # qolardi).
    mavjud = db.query(Navbat).filter(Navbat.hujjat_id == data.get("hujjatId")).first()
    if mavjud:
        db.delete(mavjud)
        # flush() - o'chirishni HOZIROQ bazaga yuboradi (lekin hali commit
        # qilinmaydi, hammasi bitta tranzaksiya bo'lib qoladi). Buni
        # qilmasak, SQLAlchemy pastdagi yangi qatorni INSERT qilishni
        # o'chirishdan OLDIN yuboradi - hujjat_id ustunidagi UNIQUE
        # cheklovga urilib, UniqueViolation beradi (sinov shuni aniq
        # ko'rsatdi).
        db.flush()
    yangi = Navbat(
        hujjat_id=data.get("hujjatId"),
        mashina_id=data.get("mashinaId"),
        raqam=data.get("raqam"),
        turi=data.get("turi"),
        shofyor=data.get("shofyor"),
        firma=data.get("firma"),
        mahsulot_id=data.get("mahsulotId"),
        mahsulot_nomi=data.get("mahsulotNomi"),
        vaqt=data.get("vaqt"),
        tuda_raqam=data.get("tudaRaqam"),
        tiket_raqam=data.get("tiketRaqam"),
        seleksiya_navi=data.get("seleksiyaNavi"),
        klass=data.get("klass"),
        sinf=data.get("sinf"),
        terim_turi=data.get("terimTuri"),
        namlik=data.get("namlik"),
        ifloslik=data.get("ifloslik"),
        tugallandi=False,
        aravalar_json=json.dumps(data.get("aravalar", {})),
    )
    db.add(yangi)
    db.commit()
    return {"status": "ok"}

@app.get("/navbat")
def navbat_get(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    from models import Navbat
    navbat = db.query(Navbat).filter(Navbat.tugallandi == False).order_by(Navbat.kelgan_vaqt.asc()).all()
    natija = []
    for n in navbat:
        natija.append({
            "hujjatId": n.hujjat_id,
            "mashinaId": n.mashina_id,
            "raqam": n.raqam,
            "turi": n.turi,
            "shofyor": n.shofyor,
            "firma": n.firma,
            "mahsulotId": n.mahsulot_id,
            "mahsulotNomi": n.mahsulot_nomi,
            "vaqt": n.vaqt,
            "tudaRaqam": n.tuda_raqam,
            "tiketRaqam": n.tiket_raqam,
            "seleksiyaNavi": n.seleksiya_navi,
            "klass": n.klass,
            "sinf": n.sinf,
            "terimTuri": n.terim_turi,
            "namlik": n.namlik,
            "ifloslik": n.ifloslik,
            "aravalar": json.loads(n.aravalar_json) if n.aravalar_json else {},
            "hujjatRaqam": db.query(Hujjat).filter(Hujjat.id == n.hujjat_id).first().raqam if n.hujjat_id else '',
        })
        
    return natija

@app.post("/navbat/tugallandi")
def navbat_tugallandi(data: dict, background_tasks: BackgroundTasks, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    from models import Navbat
    navbat = db.query(Navbat).filter(Navbat.hujjat_id == data.get("hujjatId")).first()
    if not navbat:
        raise HTTPException(status_code=404, detail="Navbat topilmadi!")

    # DIQQAT: hujjat-holat o'tishi navbat/aravalar maydonlarini
    # SAQLASHDAN OLDIN tekshiriladi va rad etilsa shu yerda 404/409
    # bilan to'xtaydi (hali db.commit() chaqirilmagani uchun bu
    # so'rovda qilingan HECH BIR o'zgarish saqlanmaydi) - shu bilan
    # operatorga "tugallandi" degan yolg'on natija hech qachon
    # qaytmaydi, na navbat, na hujjat holati yarim-yo'lda qolmaydi.
    yangi_tugallangan_hujjat_id = None
    if navbat.hujjat_id:
        hujjat = db.query(Hujjat).filter(Hujjat.id == navbat.hujjat_id).first()
        if hujjat and hujjat.holat != HujjatHolati.TUGALLANDI:
            if not holat_otishi_ruxsatmi(hujjat.holat, HujjatHolati.TUGALLANDI):
                tizim_xatosini_saqla(
                    "navbat_tugallandi",
                    f"Hujjat {hujjat.raqam} (id={hujjat.id}) '{hujjat.holat.value}' holatida edi, "
                    f"operator navbatda tugatdi, lekin avtomatik 'tugallandi'ga o'tkazilmadi "
                    f"(ruxsat etilmagan holat o'tishi)."
                )
                raise HTTPException(
                    status_code=409,
                    detail=f"Hujjat '{hujjat.holat.value}' holatida - avtomatik tugallanmaydi",
                )
            eski_holat = hujjat.holat.value
            hujjat.holat = HujjatHolati.TUGALLANDI
            db.add(TahrirTarixi(
                hujjat_id=hujjat.id,
                maydon="holat",
                eski_qiymat=eski_holat,
                yangi_qiymat=HujjatHolati.TUGALLANDI.value,
                sabab="Operator tomonidan navbatda avtomatik tugallandi",
                ozgartirgan_user_id=current_user.get("id"),
                ozgartirgan_username=current_user.get("sub"),
            ))
            # Bu - ODATIY, ENG KO'P ISHLATILADIGAN yakunlash yo'li
            # (operator "Nakladnoy chop etish"dan oldin navbatni
            # tugatganda). Excel jurnaliga aynan shu yerda, holat
            # ENDIGINA "tugallandi"ga o'tgan paytda yoziladi - shu
            # payt dostaverka/qabul_qildi/yuk_olindi kabi maydonlar
            # (odatda shu bilan bir vaqtda, alohida PUT
            # /hujjatlar/{id} orqali) allaqachon to'ldirilgan bo'ladi.
            yangi_tugallangan_hujjat_id = hujjat.id

    navbat.tugallandi = True
    navbat.tugallangan_vaqt = datetime.now()
    navbat.aravalar_json = json.dumps(data.get("aravalar", {}))

    db.commit()
    if yangi_tugallangan_hujjat_id:
        background_tasks.add_task(excel_qatorga_yoz_fon, yangi_tugallangan_hujjat_id)
    return {"status": "ok"}

@app.get("/navbat/tugallanganlar")
def tugallanganlar_get(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    from models import Navbat
    from datetime import timedelta
    kun_oldin = datetime.now() - timedelta(hours=24)
    tugallanganlar = db.query(Navbat).filter(
        Navbat.tugallandi == True,
        Navbat.tugallangan_vaqt >= kun_oldin
    ).order_by(Navbat.tugallangan_vaqt.desc()).all()
    natija = []
    for n in tugallanganlar:
        natija.append({
            "hujjatId": n.hujjat_id,
            "mashinaId": n.mashina_id,
            "raqam": n.raqam,
            "turi": n.turi,
            "shofyor": n.shofyor,
            "firma": n.firma,
            "mahsulotId": n.mahsulot_id,
            "mahsulotNomi": n.mahsulot_nomi,
            "vaqt": n.vaqt,
            "tugallanganVaqt": str(n.tugallangan_vaqt) if n.tugallangan_vaqt else None,
            "aravalar": json.loads(n.aravalar_json) if n.aravalar_json else {},
        })
    return natija

@app.post("/navbat/bekor")
def navbat_bekor(data: dict, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    from models import Navbat
    navbat = db.query(Navbat).filter(Navbat.hujjat_id == data.get("hujjatId")).first()
    if navbat:
        db.delete(navbat)
        db.commit()
    return {"status": "ok"}

@app.delete("/navbat/tozala")
def navbat_tozala(db: Session = Depends(get_db), current_user: dict = Depends(require_role("admin"))):
    from models import Navbat
    db.query(Navbat).delete()
    db.commit()
    return {"status": "ok"}

# ============ STATISTIKA ============

@app.get("/statistika/kunlik")
def kunlik_statistika(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    from datetime import date
    bugun = date.today()

    mashinalar_soni = db.query(Hujjat).filter(Hujjat.created_at >= bugun).count()
    tugallangan_soni = db.query(Hujjat).filter(
        Hujjat.holat == HujjatHolati.TUGALLANDI, Hujjat.created_at >= bugun
    ).count()
    bekor_soni = db.query(Hujjat).filter(
        Hujjat.holat == HujjatHolati.BEKOR_QILINDI, Hujjat.created_at >= bugun
    ).count()

    from models import Navbat as NavbatModel
    navbat_soni = db.query(NavbatModel).filter(NavbatModel.tugallandi == False).count()

    natijalar = db.query(
        Hujjat.mahsulot_id,
        func.count(func.distinct(Hujjat.id)).label('soni'),
        func.coalesce(func.sum(Olchov.netto), 0).label('jami_netto'),
        func.coalesce(func.sum(Olchov.konditsion), 0).label('jami_konditsion'),
    ).outerjoin(
        Olchov, Olchov.hujjat_id == Hujjat.id
    ).filter(
        Hujjat.created_at >= bugun
    ).group_by(Hujjat.mahsulot_id).all()

    natija = {}
    for row in natijalar:
        natija[row.mahsulot_id] = {
            "soni": row.soni,
            "tonnaj": round(row.jami_netto / 1000, 2),
            "konditsion": round(row.jami_konditsion / 1000, 2),
        }
    jami_tonnaj = round(sum(row.jami_netto for row in natijalar) / 1000, 2)

    bosh = {"soni": 0, "tonnaj": 0.0, "konditsion": 0.0}

    return {
        "sana": str(bugun),
        "mashinalar_soni": mashinalar_soni,
        "tugallanganlar_soni": tugallangan_soni,
        "bekor_soni": bekor_soni,
        "navbat_soni": navbat_soni,
        "chigit": natija.get(1, bosh),
        "chiganoq": {"soni": natija.get(2, bosh)["soni"], "tonnaj": natija.get(2, bosh)["tonnaj"]},
        "pochog": {"soni": natija.get(3, bosh)["soni"], "tonnaj": natija.get(3, bosh)["tonnaj"]},
        "patoz": {"soni": natija.get(4, bosh)["soni"], "tonnaj": natija.get(4, bosh)["tonnaj"]},
        "jami_tonnaj": jami_tonnaj,
    }

@app.get("/statistika/haftalik")
def haftalik_statistika(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    from datetime import date, timedelta
    bugun = date.today()
    hafta_boshi = bugun - timedelta(days=7)

    mashinalar_soni = db.query(Hujjat).filter(Hujjat.created_at >= hafta_boshi).count()
    tugallangan_soni = db.query(Hujjat).filter(
        Hujjat.holat == HujjatHolati.TUGALLANDI, Hujjat.created_at >= hafta_boshi
    ).count()
    bekor_soni = db.query(Hujjat).filter(
        Hujjat.holat == HujjatHolati.BEKOR_QILINDI, Hujjat.created_at >= hafta_boshi
    ).count()

    natijalar = db.query(
        Hujjat.mahsulot_id,
        func.count(func.distinct(Hujjat.id)).label('soni'),
        func.coalesce(func.sum(Olchov.netto), 0).label('jami_netto'),
        func.coalesce(func.sum(Olchov.konditsion), 0).label('jami_konditsion'),
    ).outerjoin(
        Olchov, Olchov.hujjat_id == Hujjat.id
    ).filter(
        Hujjat.created_at >= hafta_boshi
    ).group_by(Hujjat.mahsulot_id).all()

    natija = {}
    for row in natijalar:
        natija[row.mahsulot_id] = {
            "soni": row.soni,
            "tonnaj": round(row.jami_netto / 1000, 2),
            "konditsion": round(row.jami_konditsion / 1000, 2),
        }
    jami_tonnaj = round(sum(row.jami_netto for row in natijalar) / 1000, 2)

    bosh = {"soni": 0, "tonnaj": 0.0, "konditsion": 0.0}

    return {
        "dan": str(hafta_boshi),
        "gacha": str(bugun),
        "mashinalar_soni": mashinalar_soni,
        "tugallanganlar_soni": tugallangan_soni,
        "bekor_soni": bekor_soni,
        "chigit": natija.get(1, bosh),
        "chiganoq": {"soni": natija.get(2, bosh)["soni"], "tonnaj": natija.get(2, bosh)["tonnaj"]},
        "pochog": {"soni": natija.get(3, bosh)["soni"], "tonnaj": natija.get(3, bosh)["tonnaj"]},
        "patoz": {"soni": natija.get(4, bosh)["soni"], "tonnaj": natija.get(4, bosh)["tonnaj"]},
        "jami_tonnaj": jami_tonnaj,
    }

@app.get("/statistika/oylik")
def oylik_statistika(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    from datetime import date
    bugun = date.today()
    oy_boshi = bugun.replace(day=1)

    mashinalar_soni = db.query(Hujjat).filter(Hujjat.created_at >= oy_boshi).count()
    tugallangan_soni = db.query(Hujjat).filter(
        Hujjat.holat == HujjatHolati.TUGALLANDI, Hujjat.created_at >= oy_boshi
    ).count()
    bekor_soni = db.query(Hujjat).filter(
        Hujjat.holat == HujjatHolati.BEKOR_QILINDI, Hujjat.created_at >= oy_boshi
    ).count()

    natijalar = db.query(
        Hujjat.mahsulot_id,
        func.count(func.distinct(Hujjat.id)).label('soni'),
        func.coalesce(func.sum(Olchov.netto), 0).label('jami_netto'),
        func.coalesce(func.sum(Olchov.konditsion), 0).label('jami_konditsion'),
    ).outerjoin(
        Olchov, Olchov.hujjat_id == Hujjat.id
    ).filter(
        Hujjat.created_at >= oy_boshi
    ).group_by(Hujjat.mahsulot_id).all()

    natija = {}
    for row in natijalar:
        natija[row.mahsulot_id] = {
            "soni": row.soni,
            "tonnaj": round(row.jami_netto / 1000, 2),
            "konditsion": round(row.jami_konditsion / 1000, 2),
        }
    jami_tonnaj = round(sum(row.jami_netto for row in natijalar) / 1000, 2)

    bosh = {"soni": 0, "tonnaj": 0.0, "konditsion": 0.0}

    return {
        "oy": str(oy_boshi),
        "mashinalar_soni": mashinalar_soni,
        "tugallanganlar_soni": tugallangan_soni,
        "bekor_soni": bekor_soni,
        "chigit": natija.get(1, bosh),
        "chiganoq": {"soni": natija.get(2, bosh)["soni"], "tonnaj": natija.get(2, bosh)["tonnaj"]},
        "pochog": {"soni": natija.get(3, bosh)["soni"], "tonnaj": natija.get(3, bosh)["tonnaj"]},
        "patoz": {"soni": natija.get(4, bosh)["soni"], "tonnaj": natija.get(4, bosh)["tonnaj"]},
        "jami_tonnaj": jami_tonnaj,
    }

@app.get("/statistika/mavsum")
def mavsum_statistika(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    from datetime import date
    bugun = date.today()
    # Mavsum: 1 Avgust dan 31 Iyul gacha
    if bugun.month >= 8:
        mavsum_boshi = date(bugun.year, 8, 1)
    else:
        mavsum_boshi = date(bugun.year - 1, 8, 1)

    mashinalar_soni = db.query(Hujjat).filter(Hujjat.created_at >= mavsum_boshi).count()
    tugallangan_soni = db.query(Hujjat).filter(
        Hujjat.holat == HujjatHolati.TUGALLANDI, Hujjat.created_at >= mavsum_boshi
    ).count()
    bekor_soni = db.query(Hujjat).filter(
        Hujjat.holat == HujjatHolati.BEKOR_QILINDI, Hujjat.created_at >= mavsum_boshi
    ).count()

    natijalar = db.query(
        Hujjat.mahsulot_id,
        func.count(func.distinct(Hujjat.id)).label('soni'),
        func.coalesce(func.sum(Olchov.netto), 0).label('jami_netto'),
        func.coalesce(func.sum(Olchov.konditsion), 0).label('jami_konditsion'),
    ).outerjoin(
        Olchov, Olchov.hujjat_id == Hujjat.id
    ).filter(
        Hujjat.created_at >= mavsum_boshi
    ).group_by(Hujjat.mahsulot_id).all()

    natija = {}
    for row in natijalar:
        natija[row.mahsulot_id] = {
            "soni": row.soni,
            "tonnaj": round(row.jami_netto / 1000, 2),
            "konditsion": round(row.jami_konditsion / 1000, 2),
        }
    jami_tonnaj = round(sum(row.jami_netto for row in natijalar) / 1000, 2)

    bosh = {"soni": 0, "tonnaj": 0.0, "konditsion": 0.0}

    return {
        "mavsum_boshi": str(mavsum_boshi),
        "mashinalar_soni": mashinalar_soni,
        "tugallanganlar_soni": tugallangan_soni,
        "bekor_soni": bekor_soni,
        "chigit": natija.get(1, bosh),
        "chiganoq": {"soni": natija.get(2, bosh)["soni"], "tonnaj": natija.get(2, bosh)["tonnaj"]},
        "pochog": {"soni": natija.get(3, bosh)["soni"], "tonnaj": natija.get(3, bosh)["tonnaj"]},
        "patoz": {"soni": natija.get(4, bosh)["soni"], "tonnaj": natija.get(4, bosh)["tonnaj"]},
        "jami_tonnaj": jami_tonnaj,
    }

# ============ FIRMA / HAYDOVCHI TAHLILI ============
# Sinov paytida bazaga qolib ketgan, haqiqiy biznes faoliyati bolmagan
# firma nomlari - bular firmalar hisobotidan ATAYLAB chiqarib
# tashlanadi (foydalanuvchi tomonidan tasdiqlangan royxat). Haydovchi
# tahlilida bunday royxat yoq - chunki sinov shofyor nomlari
# (masalan "Integration Test", "Stsenariy...") turlicha va bashorat
# qilib bolmaydigan, shu sabab alohida tozalash/kelishuv talab qiladi.
_SINOV_FIRMALARI = {"Test Firma", "Sinov Firma"}


def _davr_boshlanishi(davr: str):
    from datetime import date, timedelta
    bugun = date.today()
    if davr == "kunlik":
        return bugun
    if davr == "haftalik":
        return bugun - timedelta(days=7)
    if davr == "mavsum":
        return date(bugun.year, 8, 1) if bugun.month >= 8 else date(bugun.year - 1, 8, 1)
    return bugun.replace(day=1)  # oylik (standart)


@app.get("/statistika/firmalar")
def firmalar_statistika(davr: str = "oylik", db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    """Firma boyicha jami hujjat soni/tonnaj/konditsion - eng kop
    tonnajdan boshlab saralangan. Sinov yozuvlari ("Test Firma",
    "Sinov Firma") ataylab chiqarib tashlanadi."""
    boshlanish = _davr_boshlanishi(davr)

    natijalar = db.query(
        Hujjat.firma,
        func.count(func.distinct(Hujjat.id)).label('soni'),
        func.coalesce(func.sum(Olchov.netto), 0).label('jami_netto'),
        func.coalesce(func.sum(Olchov.konditsion), 0).label('jami_konditsion'),
    ).outerjoin(
        Olchov, Olchov.hujjat_id == Hujjat.id
    ).filter(
        Hujjat.created_at >= boshlanish,
        Hujjat.holat != HujjatHolati.BEKOR_QILINDI,
        Hujjat.firma.isnot(None),
        Hujjat.firma != "",
        Hujjat.firma.notin_(_SINOV_FIRMALARI),
    ).group_by(Hujjat.firma).order_by(func.coalesce(func.sum(Olchov.netto), 0).desc()).all()

    firmalar = []
    for row in natijalar:
        jami_tonnaj = round(row.jami_netto / 1000, 2)
        jami_konditsion = round(row.jami_konditsion / 1000, 2)
        firmalar.append({
            "nom": row.firma,
            "soni": row.soni,
            "jami_tonnaj": jami_tonnaj,
            "jami_konditsion": jami_konditsion,
            "ortacha_konditsion": round(jami_konditsion / row.soni, 2) if row.soni else 0.0,
        })

    return {"davr": davr, "boshlanish": str(boshlanish), "firmalar": firmalar}


@app.get("/statistika/haydovchilar")
def haydovchilar_statistika(davr: str = "oylik", db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    """Haydovchi (shofyor) boyicha jami hujjat soni/tonnaj - eng kop
    tonnajdan boshlab saralangan."""
    boshlanish = _davr_boshlanishi(davr)

    natijalar = db.query(
        Hujjat.shofyor,
        func.count(func.distinct(Hujjat.id)).label('soni'),
        func.coalesce(func.sum(Olchov.netto), 0).label('jami_netto'),
        func.coalesce(func.sum(Olchov.konditsion), 0).label('jami_konditsion'),
    ).outerjoin(
        Olchov, Olchov.hujjat_id == Hujjat.id
    ).filter(
        Hujjat.created_at >= boshlanish,
        Hujjat.holat != HujjatHolati.BEKOR_QILINDI,
        Hujjat.shofyor.isnot(None),
        Hujjat.shofyor != "",
    ).group_by(Hujjat.shofyor).order_by(func.coalesce(func.sum(Olchov.netto), 0).desc()).all()

    haydovchilar = []
    for row in natijalar:
        jami_tonnaj = round(row.jami_netto / 1000, 2)
        jami_konditsion = round(row.jami_konditsion / 1000, 2)
        haydovchilar.append({
            "nom": row.shofyor,
            "soni": row.soni,
            "jami_tonnaj": jami_tonnaj,
            "jami_konditsion": jami_konditsion,
            "ortacha_konditsion": round(jami_konditsion / row.soni, 2) if row.soni else 0.0,
        })

    return {"davr": davr, "boshlanish": str(boshlanish), "haydovchilar": haydovchilar}

    # ============ BACKUP ============

import os

@app.post("/backup")
def backup_qilish(current_user: dict = Depends(require_role("admin"))):
    # DIQQAT (API andozasi): bu endpoint xatoda ham HAR DOIM HTTP 200
    # qaytaradi, muvaffaqiyat/xato JAVOB TANASIDAGI "status" maydoni
    # orqali bildiriladi (ko'pchilik boshqa endpoint - /login,
    # /nakladnoy/saqlash - esa to'g'ri 4xx/5xx qaytaradi). Bu ataylab -
    # frontenddagi ApiService.backupOl() shu qoidani bilib, tanani
    # tekshiradi. Agar shu endpointga YANGI chaqiruvchi qo'shsangiz,
    # FAQAT statusKod emas, albatta javob tanasidagi "status"ni ham
    # tekshiring - aks holda xato "muvaffaqiyat" deb qabul qilinishi
    # mumkin.
    try:
        backup_dir = r"C:\hazorasp_tarozi\backup"
        os.makedirs(backup_dir, exist_ok=True)

        sana = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        backup_fayl = os.path.join(backup_dir, f"backup_{sana}.sql")

        pg_dump = PG_DUMP_YOL
        db_url = make_url(DATABASE_URL)

        import subprocess
        result = subprocess.run(
            [pg_dump, "-U", db_url.username, "-p", str(db_url.port),
             "-d", db_url.database, "-f", backup_fayl],
            capture_output=True, text=True,
            env={**os.environ, "PGPASSWORD": db_url.password},
            timeout=300,
        )
        
        if result.returncode == 0:
            size = os.path.getsize(backup_fayl)
            return {
                "status": "ok",
                "fayl": backup_fayl,
                "vaqt": sana,
                "hajm": f"{size // 1024} KB"
            }
        else:
            return {"status": "error", "message": result.stderr}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/backup/royxat")
def backup_royxat(current_user: dict = Depends(require_role("admin"))):
    try:
        backup_dir = r"C:\hazorasp_tarozi\backup"
        os.makedirs(backup_dir, exist_ok=True)
        fayllar = os.listdir(backup_dir)
        fayllar.sort(reverse=True)
        return {"fayllar": fayllar, "soni": len(fayllar)}
    except Exception as e:
        return {"status": "error", "message": str(e)}
  # ============ AVTOMATIK BACKUP ============
import threading

TELEGRAM_SOZLAMA_KALIT = "oxirgi_telegram_hisobot_sanasi"

def avtomatik_telegram_hisobot():
    import time
    from datetime import date, timedelta
    from models import Sozlama
    while True:
        now = datetime.now()
        if (now.hour, now.minute) >= (8, 30):
            db = SessionLocal()
            try:
                bugun = date.today()
                kecha = bugun - timedelta(days=1)
                sozlama = db.query(Sozlama).filter(
                    Sozlama.kalit == TELEGRAM_SOZLAMA_KALIT).first()
                bugun_yuborilgan = sozlama is not None and sozlama.qiymat == str(bugun)
                if not bugun_yuborilgan:
                    mashinalar_soni = db.query(Hujjat).filter(
                        Hujjat.created_at >= kecha, Hujjat.created_at < bugun).count()

                    if kecha.month >= 8:
                        mavsum_boshi = datetime(kecha.year, 8, 1)
                    else:
                        mavsum_boshi = datetime(kecha.year - 1, 8, 1)

                    bosh3 = (0, 0.0, 0.0)

                    bugun_natijalar = db.query(
                        Hujjat.mahsulot_id,
                        func.count(func.distinct(Hujjat.id)).label('soni'),
                        func.coalesce(func.sum(Olchov.netto), 0).label('jami_netto'),
                        func.coalesce(func.sum(Olchov.konditsion), 0).label('jami_konditsion'),
                    ).outerjoin(Olchov, Olchov.hujjat_id == Hujjat.id).filter(
                        Hujjat.created_at >= kecha, Hujjat.created_at < bugun
                    ).group_by(Hujjat.mahsulot_id).all()
                    yb = {r.mahsulot_id: (r.soni, round(r.jami_netto/1000, 2), round(r.jami_konditsion/1000, 2)) for r in bugun_natijalar}
                    chigit_son, chigit_netto, chigit_kond = yb.get(1, bosh3)
                    chiganoq_son, chiganoq_netto, _ = yb.get(2, bosh3)
                    pochog_son, pochog_netto, _ = yb.get(3, bosh3)
                    patoz_son, patoz_netto, _ = yb.get(4, bosh3)

                    mavsum_natijalar = db.query(
                        Hujjat.mahsulot_id,
                        func.count(func.distinct(Hujjat.id)).label('soni'),
                        func.coalesce(func.sum(Olchov.netto), 0).label('jami_netto'),
                        func.coalesce(func.sum(Olchov.konditsion), 0).label('jami_konditsion'),
                    ).outerjoin(Olchov, Olchov.hujjat_id == Hujjat.id).filter(
                        Hujjat.created_at >= mavsum_boshi, Hujjat.created_at < bugun
                    ).group_by(Hujjat.mahsulot_id).all()
                    ym = {r.mahsulot_id: (r.soni, round(r.jami_netto/1000, 2), round(r.jami_konditsion/1000, 2)) for r in mavsum_natijalar}
                    mchigit_son, mchigit_netto, mchigit_kond = ym.get(1, bosh3)
                    mchiganoq_son, mchiganoq_netto, _ = ym.get(2, bosh3)
                    mpochog_son, mpochog_netto, _ = ym.get(3, bosh3)
                    mpatoz_son, mpatoz_netto, _ = ym.get(4, bosh3)

                    matn = f"""📊 <b>KUNLIK HISOBOT</b>
📅 Sana: {kecha}

🚛 Jami: <b>{mashinalar_soni} ta</b>

🟡 <b>Chigit:</b> {chigit_son} ta | Netto: <b>{chigit_netto} t</b> | Kond: <b>{chigit_kond} t</b>
🟢 <b>Chiganoq:</b> {chiganoq_son} ta | Netto: <b>{chiganoq_netto} t</b>
🟠 <b>Pochog:</b> {pochog_son} ta | Netto: <b>{pochog_netto} t</b>
🔴 <b>Patoz:</b> {patoz_son} ta | Netto: <b>{patoz_netto} t</b>

━━━━━━━━━━━━━━━
📦 <b>MAVSUM JAMI</b>
🟡 Chigit: {mchigit_son} ta | {mchigit_netto} t | Kond: {mchigit_kond} t
🟢 Chiganoq: {mchiganoq_son} ta | {mchiganoq_netto} t
🟠 Pochog: {mpochog_son} ta | {mpochog_netto} t
🔴 Patoz: {mpatoz_son} ta | {mpatoz_netto} t

🏭 Hazorasp Tekstil tarozi tizimi"""
                    muvaffaqiyatli = telegram_hisobot_yuborish(matn)
                    if muvaffaqiyatli:
                        if sozlama:
                            sozlama.qiymat = str(bugun)
                            sozlama.updated_at = datetime.now()
                        else:
                            db.add(Sozlama(kalit=TELEGRAM_SOZLAMA_KALIT, qiymat=str(bugun)))
                        db.commit()
                        print(f"Avtomatik hisobot yuborildi (kechagi kun: {kecha})")
                    else:
                        print("Hisobot yuborilmadi, keyingi urinishda qayta sinaladi")
            except Exception as e:
                print(f"Hisobot xato: {e}")
                tizim_xatosini_saqla("telegram_hisobot", str(e))
            finally:
                db.close()
        time.sleep(30)


BACKUP_SOZLAMA_KALIT = "oxirgi_backup_sanasi"
TARMOQ_BACKUP_SOZLAMA_KALIT = "oxirgi_tarmoq_backup_sanasi"
RASMLAR_BACKUP_SOZLAMA_KALIT = "oxirgi_rasmlar_backup_sanasi"


def tarmoqqa_backup_yubor(backup_dir: str) -> bool:
    """Lokal backup papkasini (barcha .sql fayllar) ikkinchi kompyuterdagi
    SMB ulashuvga robocopy bilan nusxalaydi. Robocopy manba/manzilda mos
    (bir xil hajm+sana) fayllarni o'tkazib yuboradi, shuning uchun butun
    papkani har safar qayta yuborish arzon - faqat yangi/etishmayotgan
    fayllar ko'chadi, hech narsa o'chirilmaydi (/MIR ishlatilmaydi)."""
    import subprocess
    unc_yol = fr"\\{TARMOQ_BACKUP_IP}\{TARMOQ_BACKUP_SHARE}"
    try:
        # Ulanish - ikkinchi komp bir muddat oldin ulangan bo'lsa ham
        # xavfsiz (Windows mavjud seansni qayta tasdiqlaydi).
        # errors="replace": net use/robocopy chiqishi tizim konsoli
        # kodировkasida (masalan cp1251) har doim ham to'g'ri
        # dekodlanavermaydi (masalan xato xabari boshqa kodировkada
        # kelsa) - shu sabab dekodlash xatosi butun urinishni
        # (returncode allaqachon to'g'ri bo'lsa ham) buzib qo'ymasin.
        subprocess.run(
            ["net", "use", unc_yol, TARMOQ_BACKUP_PAROL,
             f"/user:{TARMOQ_BACKUP_FOYDALANUVCHI}"],
            capture_output=True, text=True, errors="replace", timeout=30
        )
        natija = subprocess.run(
            ["robocopy", backup_dir, unc_yol, "*.sql", "/R:2", "/W:5"],
            capture_output=True, text=True, errors="replace", timeout=90
        )
        # Robocopy: 0-7 = muvaffaqiyat (ba'zilari hech narsa ko'chirilmadi
        # yoki manzilda ortiqcha fayl bor degani, xato emas), 8+ = xato.
        if natija.returncode >= 8:
            tizim_xatosini_saqla(
                "tarmoq_backup",
                f"robocopy xato kod bilan tugadi: {natija.returncode}")
            return False
        return True
    except Exception as e:
        tizim_xatosini_saqla("tarmoq_backup", str(e))
        return False


def rasmlar_tarmoqqa_backup_yubor() -> bool:
    """C:/RASMLAR (kamera rasmlari, nakladnoy PDF/HTML) ikkinchi
    kompyuterga zaxiralanadi - avval FAQAT baza (.sql) zaxiralanardi,
    rasmlar hech qayerga ko'chirilmasdi. Agar diskning shu qismi
    buzilsa, tortish tasdiqlovchi rasmlar qaytarib bo'lmas tarzda
    yo'qolar edi. /E - bo'sh papkalar ham (chuqur ichma-ich tuzilishni
    saqlash uchun); robocopy odatdagidek faqat yangi/o'zgargan
    fayllarni ko'chiradi - qayta-qayta chaqirish arzon."""
    import subprocess
    manba = r"C:\RASMLAR"
    if not os.path.isdir(manba):
        return True  # hali birorta rasm yo'q - qilinadigan ish yo'q
    unc_yol = fr"\\{TARMOQ_BACKUP_IP}\{TARMOQ_BACKUP_SHARE}\RASMLAR"
    try:
        subprocess.run(
            ["net", "use", fr"\\{TARMOQ_BACKUP_IP}\{TARMOQ_BACKUP_SHARE}",
             TARMOQ_BACKUP_PAROL, f"/user:{TARMOQ_BACKUP_FOYDALANUVCHI}"],
            capture_output=True, text=True, errors="replace", timeout=30
        )
        natija = subprocess.run(
            ["robocopy", manba, unc_yol, "/E", "/R:2", "/W:5"],
            capture_output=True, text=True, errors="replace", timeout=1800
        )
        if natija.returncode >= 8:
            tizim_xatosini_saqla(
                "rasmlar_tarmoq_backup",
                f"robocopy xato kod bilan tugadi: {natija.returncode}")
            return False
        return True
    except Exception as e:
        tizim_xatosini_saqla("rasmlar_tarmoq_backup", str(e))
        return False


BACKUP_RETENSIYA_KUN = 30
BACKUP_TOZALASH_SOZLAMA_KALIT = "oxirgi_backup_tozalash_sanasi"


def eski_backuplarni_tozala(backup_dir: str, kun_soni: int = BACKUP_RETENSIYA_KUN) -> int:
    """backup_dir ichidagi `kun_soni`dan eski .sql fayllarni o'chiradi.
    Avval retensiya siyosati umuman yo'q edi - har kungi backup abadiy
    jamg'arilar, disk vaqt o'tishi bilan cheksiz to'lardi. Qaytadi:
    o'chirilgan fayllar soni."""
    from datetime import timedelta
    chegara = datetime.now() - timedelta(days=kun_soni)
    ochirilgan = 0
    for nom in os.listdir(backup_dir):
        if not nom.endswith(".sql"):
            continue
        yol = os.path.join(backup_dir, nom)
        try:
            if datetime.fromtimestamp(os.path.getmtime(yol)) < chegara:
                os.remove(yol)
                ochirilgan += 1
        except OSError:
            continue
    return ochirilgan


def avtomatik_backup():
    import time
    import subprocess
    from datetime import date
    from models import Sozlama
    # Baza ulanish ma'lumotlari (foydalanuvchi/parol/port/nom) FAQAT
    # .envdagi DATABASE_URL'dan olinadi - bu yerda ikkinchi marta qo'lda
    # yozilmaydi. Aks holda parol o'zgartirilganda shu joy unutilib
    # qolib, avtomatik zaxira jimgina (faqat tizim_xatolari jadvalida)
    # ishlamay qolar edi.
    db_url = make_url(DATABASE_URL)
    tarmoq_sozlangan = bool(
        TARMOQ_BACKUP_IP and TARMOQ_BACKUP_FOYDALANUVCHI and TARMOQ_BACKUP_PAROL)
    while True:
        db = SessionLocal()
        try:
            bugun = date.today()
            backup_dir = r"C:\hazorasp_tarozi\backup"
            os.makedirs(backup_dir, exist_ok=True)

            sozlama = db.query(Sozlama).filter(
                Sozlama.kalit == BACKUP_SOZLAMA_KALIT).first()
            bugun_bajarilgan = sozlama is not None and sozlama.qiymat == str(bugun)
            if not bugun_bajarilgan:
                sana = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
                backup_fayl = os.path.join(backup_dir, f"backup_{sana}.sql")
                pg_dump = PG_DUMP_YOL
                natija = subprocess.run(
                    [pg_dump, "-U", db_url.username, "-p", str(db_url.port),
                     "-d", db_url.database, "-f", backup_fayl],
                    env={**os.environ, "PGPASSWORD": db_url.password},
                    timeout=300,
                )
                if natija.returncode == 0:
                    if sozlama:
                        sozlama.qiymat = str(bugun)
                        sozlama.updated_at = datetime.now()
                    else:
                        db.add(Sozlama(kalit=BACKUP_SOZLAMA_KALIT, qiymat=str(bugun)))
                    db.commit()
                    print(f"Avtomatik backup: {backup_fayl}")
                else:
                    xato = f"pg_dump xato kod bilan tugadi: {natija.returncode}"
                    print(f"Backup xato: {xato}")
                    tizim_xatosini_saqla("backup", xato)

            # Ikkinchi kompyuterga ko'chirish - lokal backup shu kunga
            # bajarilgan-bajarilmaganidan qat'i nazar, alohida kuzatiladi:
            # tarmoq muvaffaqiyatsiz bo'lsa ham bazani qayta dump qilish
            # shart emas, faqat ko'chirish keyingi tsikllarda qayta sinaladi.
            if tarmoq_sozlangan:
                tarmoq_sozlama = db.query(Sozlama).filter(
                    Sozlama.kalit == TARMOQ_BACKUP_SOZLAMA_KALIT).first()
                tarmoq_bugun_bajarilgan = (
                    tarmoq_sozlama is not None and tarmoq_sozlama.qiymat == str(bugun))
                if not tarmoq_bugun_bajarilgan:
                    if tarmoqqa_backup_yubor(backup_dir):
                        if tarmoq_sozlama:
                            tarmoq_sozlama.qiymat = str(bugun)
                            tarmoq_sozlama.updated_at = datetime.now()
                        else:
                            db.add(Sozlama(kalit=TARMOQ_BACKUP_SOZLAMA_KALIT, qiymat=str(bugun)))
                        db.commit()
                        print("Tarmoq backup: ikkinchi kompyuterga muvaffaqiyatli ko'chirildi")
                    else:
                        print("Tarmoq backup xato, keyingi urinishda qayta sinaladi")

                # C:/RASMLAR (kamera rasmlari, nakladnoy) - bazadan ALOHIDA
                # kuzatiladi, chunki bu ancha uzoq davom etishi mumkin
                # (/E, katta papka) - .sql backup yoki bazaning o'zi bunga
                # bog'liq bo'lib qolmasligi kerak.
                rasmlar_sozlama = db.query(Sozlama).filter(
                    Sozlama.kalit == RASMLAR_BACKUP_SOZLAMA_KALIT).first()
                rasmlar_bugun_bajarilgan = (
                    rasmlar_sozlama is not None and rasmlar_sozlama.qiymat == str(bugun))
                if not rasmlar_bugun_bajarilgan:
                    if rasmlar_tarmoqqa_backup_yubor():
                        if rasmlar_sozlama:
                            rasmlar_sozlama.qiymat = str(bugun)
                            rasmlar_sozlama.updated_at = datetime.now()
                        else:
                            db.add(Sozlama(kalit=RASMLAR_BACKUP_SOZLAMA_KALIT, qiymat=str(bugun)))
                        db.commit()
                        print("Rasmlar backup: ikkinchi kompyuterga muvaffaqiyatli ko'chirildi")
                    else:
                        print("Rasmlar backup xato, keyingi urinishda qayta sinaladi")

            # Eski backup fayllarni tozalash - kuniga bir marta yetarli.
            tozalash_sozlama = db.query(Sozlama).filter(
                Sozlama.kalit == BACKUP_TOZALASH_SOZLAMA_KALIT).first()
            tozalash_bugun_bajarilgan = (
                tozalash_sozlama is not None and tozalash_sozlama.qiymat == str(bugun))
            if not tozalash_bugun_bajarilgan:
                ochirilgan = eski_backuplarni_tozala(backup_dir)
                if ochirilgan:
                    print(f"Eski backup fayllar tozalandi: {ochirilgan} ta")
                if tozalash_sozlama:
                    tozalash_sozlama.qiymat = str(bugun)
                    tozalash_sozlama.updated_at = datetime.now()
                else:
                    db.add(Sozlama(kalit=BACKUP_TOZALASH_SOZLAMA_KALIT, qiymat=str(bugun)))
                db.commit()
        except Exception as e:
            print(f"Backup xato: {e}")
            tizim_xatosini_saqla("backup", str(e))
        finally:
            db.close()
        time.sleep(30)

# Serverni ishga tushirganda backup thread boshlash
backup_thread = threading.Thread(target=avtomatik_backup, daemon=True)
backup_thread.start()
# ============ EXCEL HISOBOT ============
import openpyxl
from openpyxl.styles import Font, PatternFill

# excel_qatorga_yoz() har chaqiruvda BUTUN faylni o'qib-qo'shib-qayta
# yozadi (openpyxl xotirada ishlaydi, qatorma-qator "qo'shish" imkonini
# bermaydi) - qulflashsiz, ikki so'rov (masalan ikki operator) bir vaqtda
# SHU BIR XIL mahsulot uchun saqlasa, ikkalasi ham faylni bir xil "eski"
# holatda o'qib olib, kim OXIRGI saqlasa o'shanikidan boshqasi butunlay
# yo'qolib qolishi mumkin edi. Har fayl yo'li uchun alohida qulf (lock)
# shu yo'qotishning oldini oladi - endi bir vaqtda kelgan so'rovlar
# navbat bilan (ketma-ket) yoziladi, hech biri boshqasini "bosib
# ketmaydi". `_excel_qulflari_royxat_lock` esa faqat YANGI qulf
# yaratishni o'zi xavfsiz (atomik) bo'lishini ta'minlaydi.
_excel_qulflari = {}
_excel_qulflari_royxat_lock = threading.Lock()


def _excel_fayl_qulfi(fayl_yol: str) -> threading.Lock:
    with _excel_qulflari_royxat_lock:
        if fayl_yol not in _excel_qulflari:
            _excel_qulflari[fayl_yol] = threading.Lock()
        return _excel_qulflari[fayl_yol]


_EXCEL_LOG_USTUNLARI_TOLIQ = [
    "№", "№ Naklad", "Mahsulot nomi", "Sana", "Tara (kg)", "Brutto (kg)",
    "Netto (kg)", "Kondicion (kg)", "Mashina raqami", "Shofyor", "Firma",
    "Tiket №", "Tuda №", "Terim turi", "Klass", "Sinf", "Seleksiya navi",
    "Namlik %", "Ifloslik %", "Dostaverka №", "Muddat", "Qabul qildi",
    "Yuk olindi",
]
# Chigit'dan boshqa 3 mahsulotda (Chiganoq, Chiganoq po'chog'i, Patoz)
# faqat asosiy ustunlar mantiqiy - Tiket/Tuda/Terim turi/Klass/Sinf/
# Seleksiya navi/Namlik/Ifloslik/Dostaverka/Muddat/Qabul qildi/Yuk olindi
# ustunlari Kondicion kabi ular uchun umuman tegishli emas, shu sabab bu
# ustunlar QATORNI EMAS, FAYLNING O'ZINI (sarlavha darajasida) chiqarib
# tashlanadi - "Excel" tugmasidagi hisobot bilan bir xil qoida.
_EXCEL_LOG_USTUNLARI_QISQA = [
    "№", "№ Naklad", "Mahsulot nomi", "Sana", "Tara (kg)", "Brutto (kg)",
    "Netto (kg)", "Mashina raqami", "Shofyor", "Firma",
]


def _oy_jurnal_malumotlarini_ol(db, mahsulot_id: int, yil: int) -> list:
    """Bitta mahsulot+yil uchun barcha (bekor qilinmagan) hujjatlarning
    to'liq ma'lumotini BATCH (N+1 so'rov muammosisiz) tarzda yig'ib
    beradi - nakladnoy_uchun_malumot() bilan bir xil mantiq (Hujjat->
    Navbat fallback zanjiri, arava-darajasida oxirgi-NULL-bo'lmagan
    qiymat, konditsion qayta hisoblash), lekin bitta hujjat o'rniga
    BUTUN yil uchun atigi 3 ta so'rovda (Hujjat, Olchov IN(...), Navbat
    IN(...)) - shu bilan excel_qatorga_yoz() mavsum davomida sekinlashib
    bormaydi (avval har hujjat uchun alohida so'rov yuborilgan
    bo'lardi)."""
    from models import Navbat

    hujjatlar = db.query(Hujjat).filter(
        Hujjat.mahsulot_id == mahsulot_id,
        Hujjat.holat != HujjatHolati.BEKOR_QILINDI,
        func.extract('year', Hujjat.created_at) == yil,
    ).order_by(Hujjat.created_at.asc()).all()
    if not hujjatlar:
        return []

    hujjat_idlar = [h.id for h in hujjatlar]

    olchovlar_dict = {}
    for o in db.query(Olchov).filter(Olchov.hujjat_id.in_(hujjat_idlar)).order_by(Olchov.id.asc()).all():
        olchovlar_dict.setdefault(o.hujjat_id, []).append(o)

    navbat_dict = {
        n.hujjat_id: n
        for n in db.query(Navbat).filter(Navbat.hujjat_id.in_(hujjat_idlar)).all()
    }

    def hn(hujjat_qiymat, navbat, maydon_nomi):
        if not _boshmi(hujjat_qiymat):
            return hujjat_qiymat
        if navbat is not None:
            navbat_qiymat = getattr(navbat, maydon_nomi)
            if not _boshmi(navbat_qiymat):
                return navbat_qiymat
        return ""

    natija = []
    for h in hujjatlar:
        navbat = navbat_dict.get(h.id)

        aravalar = {}
        for o in olchovlar_dict.get(h.id, []):
            a = aravalar.setdefault(o.arava_raqam, {
                "tara": None, "brutto": None, "netto": None,
                "namlik": None, "ifloslik": None, "konditsion": None,
            })
            for maydon in ("tara", "brutto", "netto", "namlik", "ifloslik", "konditsion"):
                qiymat = getattr(o, maydon)
                if qiymat is not None:
                    a[maydon] = qiymat
        # Konditsion HAR DOIM yakuniy (yig'ilgan) netto/namlik/ifloslikdan
        # QAYTA hisoblanadi - nakladnoy_uchun_malumot() bilan bir xil
        # qoida, izchillik uchun.
        for a in aravalar.values():
            netto = a["netto"]
            if netto is None and a["tara"] is not None and a["brutto"] is not None:
                netto = a["brutto"] - a["tara"]
            if netto is not None and a["namlik"] is not None and a["ifloslik"] is not None:
                a["konditsion"] = konditsion_hisobla(netto, a["namlik"], a["ifloslik"])
            a["netto"] = netto

        natija.append({
            "sana": h.created_at.date() if h.created_at else None,
            "raqam": h.raqam,
            "mashina_raqami": h.mashina_raqami or "",
            "shofyor": h.shofyor or "",
            "firma": h.firma or "",
            "tuda_raqam": hn(h.tuda_raqam, navbat, "tuda_raqam"),
            "tiket_raqam": hn(h.tiket_raqam, navbat, "tiket_raqam"),
            "klass": hn(h.klass, navbat, "klass"),
            "sinf": hn(h.sinf, navbat, "sinf"),
            "seleksiya_navi": hn(h.seleksiya_navi, navbat, "seleksiya_navi"),
            "terim_turi": hn(h.terim_turi, navbat, "terim_turi"),
            "namlik": (navbat.namlik if navbat and not _boshmi(navbat.namlik)
                       else aravalar.get(1, {}).get("namlik")) or "",
            "ifloslik": (navbat.ifloslik if navbat and not _boshmi(navbat.ifloslik)
                         else aravalar.get(1, {}).get("ifloslik")) or "",
            "qabul_qildi": h.qabul_qildi or "",
            "yuk_olindi": h.yuk_olindi or "",
            "dostaverka": h.dostaverka or "",
            "dostaverka_vaqt": h.dostaverka_vaqt or "",
            "aravalar": aravalar,
        })
    return natija


def excel_qatorga_yoz(hujjat_id, db):
    try:
        hujjat = db.query(Hujjat).filter(Hujjat.id == hujjat_id).first()
        if not hujjat:
            return
        # Bekor qilingan hujjatlar bu jurnalga UMUMAN yozilmaydi - "Excel"
        # tugmasidagi hisobot bilan bir xil qoida.
        if hujjat.holat == HujjatHolati.BEKOR_QILINDI:
            return
        if not hujjat.created_at:
            return

        mahsulot = db.query(Mahsulot).filter(Mahsulot.id == hujjat.mahsulot_id).first()
        if not mahsulot:
            return
        konditsiya_bormi = mahsulot.konditsiya_bor
        mahsulot_nomi = mahsulot.nom
        yil = hujjat.created_at.year

        # DIQQAT: jurnal endi APPEND emas - har safar SHU MAHSULOT+YIL
        # uchun BUTUN ma'lumot bazadan qayta yig'iladi va fayl to'liq
        # qayta quriladi (kun sarlavhalari, kunlik jami, oy oxirida
        # "OY JAMI" qatori bilan - "Excel" tugmasidagi hisobot bilan bir
        # xil, _kun_oy_guruhlab_yoz() umumiy funksiyasi orqali). Bu
        # eskirgan (guruhlanmagan) yozuvlarni ham RETROSPEKTIV to'g'ri
        # guruhlaydi - ataylab shunday (baza o'zgarmaydi, faqat hisobot
        # qayta hisoblanadi, xavfsiz).
        itemlar = _oy_jurnal_malumotlarini_ol(db, hujjat.mahsulot_id, yil)
        if not itemlar:
            return

        # Har mahsulot uchun ALOHIDA fayl (masalan hisobot_Chigit_2026.xlsx) -
        # "Excel" tugmasidagi yangi hisobot bilan bir xil mahsulot-ajratish
        # tamoyili. Eski, barcha mahsulot aralash yagona fayl
        # (C:/RASMLAR/hisobot_{yil}.xlsx) endi yangi qator OLMAYDI - u
        # tarixiy arxiv sifatida shu holicha qoladi (ataylab tegilmaydi).
        fayl_nomi_qismi = mahsulot_nomi.replace(" ", "_") or "Nomalum"
        # Papka nomi rasmlar/nakladnoy bilan AYNAN BIR XIL (probelli, asl
        # mahsulot_nomi) bo'lishi SHART - aks holda "Chiganoq po'chog'i" va
        # "Chiganoq_po'chog'i" kabi ikkita boshqa-boshqa papka hosil bo'lib,
        # bitta mahsulotning fayllari ikkiga bo'linib qoladi.
        mahsulot_papkasi = Path(f"C:/RASMLAR/{xavfsiz_papka_nomi(mahsulot_nomi)}")
        mahsulot_papkasi.mkdir(parents=True, exist_ok=True)
        fayl_yol = str(mahsulot_papkasi / f"hisobot_{fayl_nomi_qismi}_{yil}.xlsx")

        ustunlar = (
            _EXCEL_LOG_USTUNLARI_TOLIQ if konditsiya_bormi
            else _EXCEL_LOG_USTUNLARI_QISQA
        )
        ustunlar_soni = len(ustunlar)

        def _jurnal_qator_yozuvchi(ws, m, tartib_raqami):
            # Har arava uchun bitta qator - m["aravalar"] allaqachon shu
            # arava_raqam bo'yicha eng oxirgi (NULL bo'lmagan) qiymatlarga
            # jamlangan. Bitta hujjatning barcha arava-qatorlari BIR XIL
            # tartib_raqami (kun ichidagi shu hujjatning tartibi) bilan
            # chiqadi - shunda bitta mashinaning bir necha aravasi
            # jadvalda vizual ravishda birga guruhlanadi.
            netto_jami = 0
            konditsion_jami = 0
            for arava_raqam in sorted(m["aravalar"].keys()):
                a = m["aravalar"][arava_raqam]
                if a.get("tara") and a.get("brutto"):
                    netto_yaxlit = round(a["netto"]) if a.get("netto") else 0
                    konditsion_yaxlit = (
                        round(a["konditsion"])
                        if (konditsiya_bormi and a.get("konditsion")) else 0
                    )
                    asosiy = [
                        tartib_raqami,
                        m["raqam"],
                        mahsulot_nomi,
                        m["sana"].strftime("%Y-%m-%d"),
                        round(a["tara"]),
                        round(a["brutto"]),
                        netto_yaxlit if a.get("netto") else "",
                    ]
                    if konditsiya_bormi:
                        qator = asosiy + [
                            konditsion_yaxlit if a.get("konditsion") else "",
                            m["mashina_raqami"],
                            m["shofyor"],
                            m["firma"],
                            m["tiket_raqam"],
                            m["tuda_raqam"],
                            m["terim_turi"],
                            m["klass"],
                            m["sinf"],
                            m["seleksiya_navi"],
                            m["namlik"],
                            m["ifloslik"],
                            m["dostaverka"],
                            m["dostaverka_vaqt"],
                            m["qabul_qildi"],
                            m["yuk_olindi"],
                        ]
                    else:
                        qator = asosiy + [
                            m["mashina_raqami"],
                            m["shofyor"],
                            m["firma"],
                        ]
                    ws.append(qator)
                    netto_jami += netto_yaxlit
                    konditsion_jami += konditsion_yaxlit
            return netto_jami, konditsion_jami

        # O'qish-o'zgartirish-yozish (read-modify-write) bo'linmas
        # (atomik) bo'lishi uchun shu FAYLGA xos qulf ostida bajariladi -
        # boshqa mahsulot fayliga yozish bilan hech qachon to'sqinlik
        # qilmaydi, faqat AYNAN shu faylga bir vaqtda yozilishi mumkin
        # bo'lgan so'rovlar navbatga turadi.
        with _excel_fayl_qulfi(fayl_yol):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Hisobot"

            _kun_oy_guruhlab_yoz(
                ws, itemlar, konditsiya_bormi, _jurnal_qator_yozuvchi,
                ustunlar_soni=ustunlar_soni,
                sarlavha_ustun=3, netto_ustun=7,
                konditsion_ustun=8 if konditsiya_bormi else None,
                ustunlar_royxati=ustunlar,
            )

            wb.save(fayl_yol)
        print(f"Excel jurnal qayta qurildi: {mahsulot_nomi} ({yil}-yil, {len(itemlar)} hujjat)")
    except Exception as e:
        print(f"Excel xato: {e}")
        tizim_xatosini_saqla("excel", f"{mahsulot_nomi} ({yil}-yil): {e}")


def excel_qatorga_yoz_fon(hujjat_id):
    """excel_qatorga_yoz()ni FON vazifasi sifatida chaqirish uchun
    o'ram - operator "Saqlash" javobini kutib turmasligi uchun (mavsum
    davomida yillik hujjatlar ko'payishi bilan bitta chaqiruv bir necha
    soniya olishi mumkin edi, buni endi operator SEZMAYDI). So'rovning
    o'z DB sessiyasi ISHLATILMAYDI (javob jo'natilgach yopiladi) -
    o'rniga bu yerda ALOHIDA, yangi sessiya ochiladi/yopiladi, xuddi
    avtomatik_backup()/avtomatik_telegram_hisobot() fon oqimlaridagi
    naqsh bilan bir xil. excel_qatorga_yoz()ning o'zi butun tanasini
    try/except bilan o'rab, hech qachon xato otmaydi (qarang: yuqorida)
    - shu sabab bu yerda qo'shimcha himoya shart emas."""
    db = SessionLocal()
    try:
        excel_qatorga_yoz(hujjat_id, db)
    finally:
        db.close()

# ============ SOZLAMALAR ============

@app.get("/sozlamalar")
def sozlamalar_olish(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    sozlamalar = db.query(Sozlama).all()
    natija = {}
    for s in sozlamalar:
        natija[s.kalit] = s.qiymat
    # Operator ekrani ham shu endpointdan (tuda_raqam/klass/sinf va
    # h.k. standart qiymatlar uchun) foydalanadi, shu sabab butun
    # endpoint admin-only qilinmaydi - faqat sezgir maydon (Telegram bot
    # tokeni) admin/hisobchi bo'lmagan rollardan yashiriladi.
    if current_user.get("role") not in ("admin", "hisobchi"):
        natija.pop("telegram_token", None)
    # Moliyaviy PIN xeshi HECH QACHON (hatto admin/hisobchiga ham) shu
    # umumiy endpoint orqali qaytarilmaydi - buni faqat maxsus
    # /moliyaviy/pin-tekshir orqali (xeshning o'zi emas, faqat
    # to'g'ri/xato natija) bilish mumkin.
    natija.pop(MOLIYAVIY_PIN_SOZLAMA_KALIT, None)
    return natija

@app.post("/sozlamalar")
def sozlama_saqlash(data: dict, db: Session = Depends(get_db), current_user: dict = Depends(require_role("admin"))):
    # Moliyaviy PIN shu umumiy endpoint orqali O'RNATILISHI/O'ZGARTIRILISHI
    # MUMKIN EMAS - aks holda /moliyaviy/pin'dagi "eski PIN talab qilinadi"
    # himoyasi butunlay chetlab o'tilardi (bu yerda hech qanday eski PIN
    # tekshiruvi yo'q). Faqat PUT /moliyaviy/pin orqali o'zgartiriladi.
    if MOLIYAVIY_PIN_SOZLAMA_KALIT in data:
        raise HTTPException(
            status_code=400,
            detail="Moliyaviy PIN faqat /moliyaviy/pin orqali o'zgartiriladi!",
        )
    for kalit, qiymat in data.items():
        mavjud = db.query(Sozlama).filter(Sozlama.kalit == kalit).first()
        if mavjud:
            mavjud.qiymat = str(qiymat)
            mavjud.updated_at = datetime.now()
        else:
            yangi = Sozlama(kalit=kalit, qiymat=str(qiymat))
            db.add(yangi)
    db.commit()
    return {"status": "ok"}

# ============ TIZIM XATOLARI RO'YXATI ============

@app.get("/tizim-xatolari")
def tizim_xatolari_royxati(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in ("admin", "hisobchi"):
        raise HTTPException(status_code=403, detail="Bu amal uchun sizda ruxsat yo'q!")
    xatolar = db.query(TizimXatosi).order_by(
        TizimXatosi.created_at.desc()
    ).limit(50).all()
    return [
        {
            "id": x.id,
            "turi": x.turi,
            "xabar": x.xabar,
            "korilgan": x.korilgan,
            "vaqt": str(x.created_at),
        }
        for x in xatolar
    ]

@app.post("/tizim-xatolari/{xato_id}/korildi")
def tizim_xatosi_korildi(xato_id: int, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in ("admin", "hisobchi"):
        raise HTTPException(status_code=403, detail="Bu amal uchun sizda ruxsat yo'q!")
    xato = db.query(TizimXatosi).filter(TizimXatosi.id == xato_id).first()
    if xato:
        xato.korilgan = True
        db.commit()
    return {"status": "ok"}

# ============ SERVER HOLATI ============
import psutil

@app.get("/server/holat")
def server_holat(current_user: dict = Depends(get_current_user)):
    try:
        cpu = psutil.cpu_percent(interval=1)
        ram = psutil.virtual_memory()
        disk = psutil.disk_usage('C:/')
        uptime_seconds = (datetime.now() - datetime.fromtimestamp(psutil.boot_time())).total_seconds()
        kun = int(uptime_seconds // 86400)
        soat = int((uptime_seconds % 86400) // 3600)
        return {
            "cpu": round(cpu, 1),
            "ram": round(ram.percent, 1),
            "disk": round(disk.percent, 1),
            "uptime": f"{kun} kun {soat} soat"
        }
    except Exception as e:
        # Avval bu yerda hech narsa log qilinmasdi - monitoring
        # endpoint'ining o'zi buzilib qolsa, dashboard "hammasi 0%,
        # tinch" deb noto'g'ri ko'rsatib turaverar edi.
        tizim_xatosini_saqla("server_holat", str(e))
        return {"cpu": 0, "ram": 0, "disk": 0, "uptime": "—"}

# ============ TIZIM XATOLARI ============

def tizim_xatosini_saqla(turi: str, xabar: str):
    db = SessionLocal()
    try:
        yangi = TizimXatosi(turi=turi, xabar=xabar)
        db.add(yangi)
        db.commit()
    except Exception as e:
        print(f"Tizim xatosini saqlashda xato: {e}")
    finally:
        db.close()

# ============ TELEGRAM BOT ============
import requests as req

def telegram_xabar_yuborish(matn: str) -> bool:
    """Ogohlantirish/xatolik boti - tarozi, tunnel, kamera va shunga
    o'xshash texnik muammolar shu yerdan yuboriladi. Kunlik hisobotlar
    UCHUN EMAS - qarang: telegram_hisobot_yuborish()."""
    try:
        from config import TELEGRAM_TOKEN, TELEGRAM_CHAT_ID
        if not TELEGRAM_TOKEN or not TELEGRAM_CHAT_ID:
            return False
        url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
        javob = req.post(url, json={
            "chat_id": TELEGRAM_CHAT_ID,
            "text": matn,
            "parse_mode": "HTML"
        }, timeout=5)
        javob.raise_for_status()
        return True
    except Exception as e:
        print(f"Telegram xato: {e}")
        tizim_xatosini_saqla("telegram", str(e))
        return False


def telegram_hisobot_yuborish(matn: str) -> bool:
    """Statistika/hisobot boti - kunlik ishlab chiqarish hisoboti shu
    yerdan yuboriladi. ALOHIDA bot/chat - ogohlantirish xabarlari bilan
    aralashmasligi uchun (masalan rahbar/buxgalter shu botga qo'shilishi
    mumkin, texnik xatolik xabarlarini ko'rmasdan). Hali sozlanmagan
    bo'lsa (token/chat_id bo'sh) - xato bermay, jimgina False qaytaradi."""
    try:
        from config import TELEGRAM_HISOBOT_TOKEN, TELEGRAM_HISOBOT_CHAT_ID
        if not TELEGRAM_HISOBOT_TOKEN or not TELEGRAM_HISOBOT_CHAT_ID:
            return False
        url = f"https://api.telegram.org/bot{TELEGRAM_HISOBOT_TOKEN}/sendMessage"
        javob = req.post(url, json={
            "chat_id": TELEGRAM_HISOBOT_CHAT_ID,
            "text": matn,
            "parse_mode": "HTML"
        }, timeout=5)
        javob.raise_for_status()
        return True
    except Exception as e:
        print(f"Telegram hisobot xato: {e}")
        tizim_xatosini_saqla("telegram_hisobot", str(e))
        return False

hisobot_thread = threading.Thread(target=avtomatik_telegram_hisobot, daemon=True)
hisobot_thread.start()

# ============ TUNNEL/SERVER O'Z-O'ZINI KUZATISH ============
# Ilgari hech narsa serverning/tunnel'ning o'zi ishlamay qolganini
# kuzatmasdi - faqat operator ilovani ochishga urinib, ishlamasligini
# payqashi orqaligina ma'lum bo'lardi. Bu fon oqim davriy ravishda
# serverning O'Z ommaviy manziliga (SERVER_ASOSIY_URL) so'rov yuborib,
# butun zanjir (backend -> Cloudflare Tunnel -> Cloudflare edge)
# ishlab turganini tekshiradi.
_TUNNEL_TEKSHIRUV_OSIYA = 180
_TUNNEL_XATO_CHEGARA = 2

_tunnel_holati = {"ketma_ket": 0, "ogohlantirilgan": False}
_tunnel_holati_qulf = threading.Lock()


def _tunnel_bir_tekshiruv():
    """Bitta tekshiruv sikli - alohida funksiya qilib ajratilgan, shunda
    sinovlarda `while True`/`time.sleep`ga tegmasdan to'g'ridan-to'g'ri
    chaqirish mumkin."""
    try:
        javob = req.get(f"{SERVER_ASOSIY_URL}/health", timeout=10)
        muvaffaqiyat = javob.status_code == 200
    except Exception:
        muvaffaqiyat = False

    yuboriladigan_matn = None
    with _tunnel_holati_qulf:
        if not muvaffaqiyat:
            _tunnel_holati["ketma_ket"] += 1
            if (_tunnel_holati["ketma_ket"] >= _TUNNEL_XATO_CHEGARA
                    and not _tunnel_holati["ogohlantirilgan"]):
                _tunnel_holati["ogohlantirilgan"] = True
                yuboriladigan_matn = (
                    f"🔴 <b>Diqqat!</b> Server o'zining ommaviy manziliga "
                    f"({html.escape(SERVER_ASOSIY_URL)}) yeta olmayapti - "
                    f"Cloudflare Tunnel yoki tarmoq muammosi bo'lishi mumkin."
                )
        else:
            if _tunnel_holati["ogohlantirilgan"]:
                yuboriladigan_matn = "✅ Server ommaviy manzili qayta ishlay boshladi."
            _tunnel_holati["ketma_ket"] = 0
            _tunnel_holati["ogohlantirilgan"] = False

    if yuboriladigan_matn:
        telegram_xabar_yuborish(yuboriladigan_matn)


def _tunnel_ozini_tekshirish():
    while True:
        time.sleep(_TUNNEL_TEKSHIRUV_OSIYA)
        _tunnel_bir_tekshiruv()


# ============ TAROZI AGENTI O'Z-O'ZINI KUZATISH (Telegram) ============
# _tarozi_oxirgi_yangilanish (yuqorida, TAROZI bo'limida) hozirgacha
# FAQAT operator ekrani GET /tarozi/joriy so'raganda tekshirilardi - agar
# hech kim ekranni ochmasa (masalan tunda), agent uzilib qolsa ham hech
# kim bilmasdi. Bu fon oqim buni davriy ravishda, MUSTAQIL tekshiradi.
# DIQQAT: bu FAQAT backend TIRIK bo'lganda ishlaydi - agar backend
# jarayonining o'zi butunlay o'lsa, bu thread ham u bilan birga o'ladi
# (shu sabab tarozi_agent.py'ning o'zida ham ALOHIDA, mustaqil ogohlantirish
# bor - qarang: tarozi_agent/tarozi_agent.py).
_TAROZI_ALERT_TEKSHIRUV_OSIYA = 60
# Operator ekranidagi "Uzilgan" belgisi 10 soniyada ko'rinadi (tezkor,
# vizual) - lekin Telegram ogohlantirish uchun ancha yuqoriroq chegara
# ishlatiladi, aks holda har bir qisqa (bir necha soniyalik) uzilishda
# ham xabar kelib, spam bo'lib qolardi.
_TAROZI_ALERT_CHEGARA_SONIYA = 180

_tarozi_alert_holati = {"ogohlantirilgan": False}
_tarozi_alert_holati_qulf = threading.Lock()


def _tarozi_alert_bir_tekshiruv():
    with _tarozi_qulf:
        oxirgi = _tarozi_oxirgi_yangilanish
    muddat_otdi = (time.time() - oxirgi) if oxirgi else float("inf")
    uzilgan = muddat_otdi > _TAROZI_ALERT_CHEGARA_SONIYA

    yuboriladigan_matn = None
    with _tarozi_alert_holati_qulf:
        if uzilgan:
            if not _tarozi_alert_holati["ogohlantirilgan"]:
                _tarozi_alert_holati["ogohlantirilgan"] = True
                # muddat_otdi cheksiz bo'lishi mumkin (server hozirgina
                # ishga tushgan, agentdan HALI BIRON MARTA ham ma'lumot
                # kelmagan) - int(inf) OverflowError beradi, shu sabab
                # bu holat alohida xabar bilan ko'rsatiladi.
                muddat_matni = (
                    "hali birorta ham ma'lumot kelmagan"
                    if muddat_otdi == float("inf")
                    else f"{int(muddat_otdi)} soniyadan buyon ma'lumot kelmayapti"
                )
                yuboriladigan_matn = (
                    f"🔴 <b>Diqqat!</b> Tarozi agentidan {muddat_matni} - "
                    f"TaroziAgent yoki tarozixona kompyuteri/tarmog'i bilan "
                    f"muammo bo'lishi mumkin."
                )
        else:
            if _tarozi_alert_holati["ogohlantirilgan"]:
                yuboriladigan_matn = "✅ Tarozi agentidan ma'lumot qayta kela boshladi."
            _tarozi_alert_holati["ogohlantirilgan"] = False

    if yuboriladigan_matn:
        telegram_xabar_yuborish(yuboriladigan_matn)


def _tarozi_alert_kuzatuvchisi():
    while True:
        time.sleep(_TAROZI_ALERT_TEKSHIRUV_OSIYA)
        _tarozi_alert_bir_tekshiruv()


tarozi_alert_thread = threading.Thread(target=_tarozi_alert_kuzatuvchisi, daemon=True)
tarozi_alert_thread.start()

tunnel_kuzatuv_thread = threading.Thread(target=_tunnel_ozini_tekshirish, daemon=True)
tunnel_kuzatuv_thread.start()

# ============ CLOUDFLAREDTUNNEL WINDOWS XIZMATINI KUZATISH (Telegram) ============
# Yuqoridagi _tunnel_ozini_tekshirish() serverning ommaviy manziliga HTTP
# so'rov yuborib, butun zanjirni (backend+tunnel+Cloudflare edge) birgalikda
# tekshiradi - lekin aynan QAYSI qism buzilganini aytmaydi. Bu fon oqim
# to'g'ridan-to'g'ri, tarmoq orqali EMAS, shu kompyuterning o'zida
# CloudflaredTunnel Windows xizmatining holatini tekshiradi - shu sabab
# tezroq va ANIQROQ ("aynan tunnel xizmati to'xtadi") ogohlantirish beradi.
# DIQQAT: bu ham backend TIRIK bo'lgandagina ishlaydi - agar backend
# jarayonining o'zi (yoki butun kompyuter) o'lsa, TaroziAgent'dagi mustaqil
# Qatlam 2 baribir "serverga ulanib bo'lmayapti" deb ogohlantiradi (u ham
# xuddi shu CloudflaredTunnel orqali ulanadi) - shu sabab bu yerga
# UCHINCHI, alohida jarayon qo'shishning hojati yo'q.
_TUNNEL_XIZMAT_TEKSHIRUV_OSIYA = 60
_CLOUDFLARE_XIZMAT_NOMI = "CloudflaredTunnel"

_tunnel_xizmat_holati = {"ogohlantirilgan": False}
_tunnel_xizmat_holati_qulf = threading.Lock()


def _tunnel_xizmat_ishlab_turibdimi():
    import subprocess
    try:
        natija = subprocess.run(
            ["sc", "query", _CLOUDFLARE_XIZMAT_NOMI],
            capture_output=True, text=True, timeout=10,
        )
        return "RUNNING" in natija.stdout
    except Exception:
        return False


def _tunnel_xizmat_bir_tekshiruv():
    ishlayaptimi = _tunnel_xizmat_ishlab_turibdimi()

    yuboriladigan_matn = None
    with _tunnel_xizmat_holati_qulf:
        if not ishlayaptimi:
            if not _tunnel_xizmat_holati["ogohlantirilgan"]:
                _tunnel_xizmat_holati["ogohlantirilgan"] = True
                yuboriladigan_matn = (
                    "🔴 <b>Diqqat!</b> CloudflaredTunnel Windows xizmati "
                    "to'xtab qoldi - tashqi (jamoat) kirish (smart-tarozi.uz) "
                    "ishlamasligi mumkin. Windows xizmatini tekshiring."
                )
        else:
            if _tunnel_xizmat_holati["ogohlantirilgan"]:
                yuboriladigan_matn = "✅ CloudflaredTunnel xizmati qayta ishga tushdi."
            _tunnel_xizmat_holati["ogohlantirilgan"] = False

    if yuboriladigan_matn:
        telegram_xabar_yuborish(yuboriladigan_matn)


def _tunnel_xizmat_kuzatuvchisi():
    while True:
        time.sleep(_TUNNEL_XIZMAT_TEKSHIRUV_OSIYA)
        _tunnel_xizmat_bir_tekshiruv()


tunnel_xizmat_thread = threading.Thread(target=_tunnel_xizmat_kuzatuvchisi, daemon=True)
tunnel_xizmat_thread.start()

@app.post("/telegram/test")
def telegram_test(current_user: dict = Depends(require_role("admin"))):
    telegram_xabar_yuborish("✅ Hazorasp Tekstil tarozi tizimi ulandi!")
    return {"status": "ok"}

@app.get("/telegram/kunlik")
def telegram_kunlik(db: Session = Depends(get_db), current_user: dict = Depends(require_role("admin"))):
    from datetime import date
    bugun = date.today()
    mashinalar_soni = db.query(Hujjat).filter(Hujjat.created_at >= bugun).count()

    if bugun.month >= 8:
        mavsum_boshi = datetime(bugun.year, 8, 1)
    else:
        mavsum_boshi = datetime(bugun.year - 1, 8, 1)

    bosh3 = (0, 0.0, 0.0)

    bugun_natijalar = db.query(
        Hujjat.mahsulot_id,
        func.count(func.distinct(Hujjat.id)).label('soni'),
        func.coalesce(func.sum(Olchov.netto), 0).label('jami_netto'),
        func.coalesce(func.sum(Olchov.konditsion), 0).label('jami_konditsion'),
    ).outerjoin(Olchov, Olchov.hujjat_id == Hujjat.id).filter(
        Hujjat.created_at >= bugun
    ).group_by(Hujjat.mahsulot_id).all()
    yb = {r.mahsulot_id: (r.soni, round(r.jami_netto/1000, 2), round(r.jami_konditsion/1000, 2)) for r in bugun_natijalar}
    chigit_son, chigit_netto, chigit_kond = yb.get(1, bosh3)
    chiganoq_son, chiganoq_netto, _ = yb.get(2, bosh3)
    pochog_son, pochog_netto, _ = yb.get(3, bosh3)
    patoz_son, patoz_netto, _ = yb.get(4, bosh3)

    mavsum_natijalar = db.query(
        Hujjat.mahsulot_id,
        func.count(func.distinct(Hujjat.id)).label('soni'),
        func.coalesce(func.sum(Olchov.netto), 0).label('jami_netto'),
        func.coalesce(func.sum(Olchov.konditsion), 0).label('jami_konditsion'),
    ).outerjoin(Olchov, Olchov.hujjat_id == Hujjat.id).filter(
        Hujjat.created_at >= mavsum_boshi
    ).group_by(Hujjat.mahsulot_id).all()
    ym = {r.mahsulot_id: (r.soni, round(r.jami_netto/1000, 2), round(r.jami_konditsion/1000, 2)) for r in mavsum_natijalar}
    mchigit_son, mchigit_netto, mchigit_kond = ym.get(1, bosh3)
    mchiganoq_son, mchiganoq_netto, _ = ym.get(2, bosh3)
    mpochog_son, mpochog_netto, _ = ym.get(3, bosh3)
    mpatoz_son, mpatoz_netto, _ = ym.get(4, bosh3)

    matn = f"""📊 <b>KUNLIK HISOBOT</b>
📅 Sana: {bugun}

🚛 Jami: <b>{mashinalar_soni} ta</b>

🟡 <b>Chigit:</b> {chigit_son} ta | Netto: <b>{chigit_netto} t</b> | Kond: <b>{chigit_kond} t</b>
🟢 <b>Chiganoq:</b> {chiganoq_son} ta | Netto: <b>{chiganoq_netto} t</b>
🟠 <b>Pochog:</b> {pochog_son} ta | Netto: <b>{pochog_netto} t</b>
🔴 <b>Patoz:</b> {patoz_son} ta | Netto: <b>{patoz_netto} t</b>

━━━━━━━━━━━━━━━
📦 <b>MAVSUM JAMI</b>
🟡 Chigit: {mchigit_son} ta | {mchigit_netto} t | Kond: {mchigit_kond} t
🟢 Chiganoq: {mchiganoq_son} ta | {mchiganoq_netto} t
🟠 Pochog: {mpochog_son} ta | {mpochog_netto} t
🔴 Patoz: {mpatoz_son} ta | {mpatoz_netto} t

🏭 Hazorasp Tekstil tarozi tizimi"""
    
    telegram_hisobot_yuborish(matn)
    return {"status": "ok", "xabar": matn}
 # ============ PDF SAQLASH ============

def _boshmi(qiymat):
    """None va bo'sh satrni 'qiymat yo'q' deb hisoblaydi - fallback zanjiri uchun."""
    if qiymat is None:
        return True
    if isinstance(qiymat, str) and qiymat.strip() == "":
        return True
    return False


def nakladnoy_uchun_malumot(db: Session, hujjat_id: int) -> dict:
    """Nakladnoy PDF uchun kerakli barcha ma'lumotni bazadan to'liq, ishonchli
    yig'ib beradi - frontend ekran holatiga (matn controller'lariga) bog'liq emas.

    Manba tartibi:
    - mashina_raqami/shofyor/firma/raqam/aravalar_soni - Hujjat'dan
    - mashina_turi - Mashina'dan (mashina_id orqali)
    - mahsulot_nomi - Mahsulot'dan (mahsulot_id orqali)
    - tiket_raqam/klass/sinf/seleksiya_navi/terim_turi/tuda_raqam - avval Hujjat'ning
      o'z ustuni, bo'sh bo'lsa Navbat'dagi mos ustunga fallback (eski hujjatlarda bu
      ustunlar Hujjat'ga backfill qilingan, yangilarida hali faqat Navbat'da bor)
    - qabul_qildi/yuk_olindi/dostaverka/dostaverka_vaqt - faqat Hujjat'dan (Navbat'da
      bunday ustunlar yo'q)
    - har arava uchun tara/brutto/netto/namlik/ifloslik/konditsion - Olchov'dan,
      shu arava_raqam bo'yicha eng oxirgi ma'lum (NULL bo'lmagan) qiymatlar
    """
    hujjat = db.query(Hujjat).filter(Hujjat.id == hujjat_id).first()
    if not hujjat:
        return None

    mashina = db.query(Mashina).filter(Mashina.id == hujjat.mashina_id).first()
    mahsulot = db.query(Mahsulot).filter(Mahsulot.id == hujjat.mahsulot_id).first()

    from models import Navbat
    navbat = db.query(Navbat).filter(Navbat.hujjat_id == hujjat_id).first()

    def hn(hujjat_maydon, navbat_maydon):
        """Hujjat -> Navbat fallback zanjiri."""
        if not _boshmi(hujjat_maydon):
            return hujjat_maydon
        if navbat is not None and not _boshmi(navbat_maydon):
            return navbat_maydon
        return ""

    # Har arava_raqam uchun barcha Olchov qatorlarini id bo'yicha o'sish tartibida
    # ko'rib chiqib, har bir maydonni faqat YANGI qiymat NULL bo'lmaganda yangilaymiz -
    # shunda keyingi (masalan faqat brutto yozilgan) qator avvalgi tara qiymatini
    # "yo'qotib qo'ymaydi".
    olchov_qatorlari = db.query(Olchov).filter(
        Olchov.hujjat_id == hujjat_id
    ).order_by(Olchov.id.asc()).all()

    aravalar = {}
    for o in olchov_qatorlari:
        a = aravalar.setdefault(o.arava_raqam, {
            "tara": None, "brutto": None, "netto": None,
            "namlik": None, "ifloslik": None, "konditsion": None,
        })
        for maydon in ("tara", "brutto", "netto", "namlik", "ifloslik", "konditsion"):
            qiymat = getattr(o, maydon)
            if qiymat is not None:
                a[maydon] = qiymat

    # DIQQAT (vaqt-bog'liq xato tuzatildi): yuqoridagi tsikl netto/namlik/
    # ifloslik/konditsion'ni BIR-BIRIDAN MUSTAQIL ravishda ("oxirgi
    # NULL-bo'lmagan qiymat") oladi - agar bu maydonlar turli Olchov
    # qatorlarida turli vaqtda yozilgan bo'lsa (masalan keyinroq faqat
    # namlik/ifloslik tuzatilsa), saqlangan "konditsion" ESKI, mos
    # kelmaydigan qiymatda qolib ketishi mumkin edi. Shu sabab bu yerda
    # konditsion HAR DOIM yakuniy (yig'ilgan) netto/namlik/ifloslikdan
    # QAYTA hisoblanadi - Nakladnoy (va shu funksiyaga tayanadigan barcha
    # boshqa joy: QR-ko'rish, Excel jurnal) doim izchil qiymat ko'rsatadi.
    for a in aravalar.values():
        netto = a["netto"]
        if netto is None and a["tara"] is not None and a["brutto"] is not None:
            netto = a["brutto"] - a["tara"]
        if netto is not None and a["namlik"] is not None and a["ifloslik"] is not None:
            a["konditsion"] = konditsion_hisobla(netto, a["namlik"], a["ifloslik"])

    # Kirgan vaqt = tara BIRINCHI o'lchangan payt, chiqqan vaqt = brutto
    # OXIRGI o'lchangan payt - Navbat qatoriga bog'liq emas (ba'zi
    # hujjatlarda Navbat qatori umuman bo'lmasligi mumkin), shu sabab
    # to'g'ridan-to'g'ri Olchov.created_at'dan hisoblanadi.
    tara_vaqtlari = [o.created_at for o in olchov_qatorlari if o.tara is not None]
    brutto_vaqtlari = [o.created_at for o in olchov_qatorlari if o.brutto is not None]
    kirgan_vaqt = min(tara_vaqtlari) if tara_vaqtlari else None
    chiqqan_vaqt = max(brutto_vaqtlari) if brutto_vaqtlari else None

    return {
        "hujjat_id": hujjat.id,
        "raqam": hujjat.raqam,
        "sana": hujjat.created_at.strftime("%Y-%m-%d") if hujjat.created_at else "",
        "kirgan_vaqt": kirgan_vaqt.strftime("%Y-%m-%d %H:%M:%S") if kirgan_vaqt else "",
        "chiqqan_vaqt": chiqqan_vaqt.strftime("%Y-%m-%d %H:%M:%S") if chiqqan_vaqt else "",
        "mashina_raqami": hujjat.mashina_raqami or "",
        "mashina_turi": mashina.turi if mashina else "",
        "shofyor": hujjat.shofyor or "",
        "firma": hujjat.firma or "",
        "mahsulot_nomi": mahsulot.nom if mahsulot else "",
        "aravalar_soni": hujjat.aravalar_soni or 1,
        "tuda_raqam": hn(hujjat.tuda_raqam, navbat.tuda_raqam if navbat else None),
        "tiket_raqam": hn(hujjat.tiket_raqam, navbat.tiket_raqam if navbat else None),
        "klass": hn(hujjat.klass, navbat.klass if navbat else None),
        "sinf": hn(hujjat.sinf, navbat.sinf if navbat else None),
        "seleksiya_navi": hn(hujjat.seleksiya_navi, navbat.seleksiya_navi if navbat else None),
        "terim_turi": hn(hujjat.terim_turi, navbat.terim_turi if navbat else None),
        # Hujjat'da namlik/ifloslik ustuni yo'q (faqat Navbat va Olchov'da bor) -
        # avval Navbat'ga, u ham bo'lmasa 1-aravaning Olchov qiymatiga qaraladi.
        "namlik": (navbat.namlik if navbat and not _boshmi(navbat.namlik)
                   else aravalar.get(1, {}).get("namlik")) or "",
        "ifloslik": (navbat.ifloslik if navbat and not _boshmi(navbat.ifloslik)
                     else aravalar.get(1, {}).get("ifloslik")) or "",
        "qabul_qildi": hujjat.qabul_qildi or "",
        "yuk_olindi": hujjat.yuk_olindi or "",
        "dostaverka": hujjat.dostaverka or "",
        "dostaverka_vaqt": hujjat.dostaverka_vaqt or "",
        "aravalar": aravalar,
    }


_NAKLADNOY_STIL = """
* { box-sizing: border-box; }
body { font-family: Arial, sans-serif; font-size: 13px; margin: 0; padding: 0; background: white; color: #0D1B2A; }
.sahifa { padding: 10mm 12mm; position: relative; }
.nusxa-badge { background: #0D1B2A; color: white; text-align: center; padding: 6px; border-radius: 6px; font-size: 12px; font-weight: 700; letter-spacing: 1px; margin-bottom: 8px; }
.sarlavha { text-align: center; font-size: 16px; font-weight: 700; color: #0F6E56; margin: 4px 0 2px 0; }
.subtitle { text-align: center; font-size: 11px; color: #607080; margin-bottom: 10px; }
.karta { background: white; border: 1px solid #D8EDD0; border-radius: 8px; padding: 8px 14px; margin-bottom: 8px; }
.maydon { display: flex; justify-content: space-between; padding: 3px 0; font-size: 12px; }
.label { color: #607080; }
.qiymat { font-weight: 600; text-align: right; }
.info-grid { display: grid; grid-template-columns: repeat(4, 1fr); gap: 8px 16px; }
table { width: 100%; border-collapse: collapse; margin-top: 4px; }
th { background: #0F6E56; color: white; padding: 8px 6px; font-size: 11px; }
td { border: 1px solid #D8EDD0; padding: 8px 6px; text-align: center; font-size: 12px; }
.jami td { font-weight: 700; background: #EAF6F0; }
.dostaverka-box { background: #EAF6F0; border: 1px solid #B8E0C8; border-radius: 6px; padding: 8px 12px; margin-top: 8px; display: flex; justify-content: space-between; font-size: 12px; }
.imzo-grid { display: grid; grid-template-columns: repeat(4, 1fr) 70px; gap: 14px; margin-top: 18px; align-items: end; font-size: 10px; color: #607080; text-align: center; }
.imzo-line { border-bottom: 1px solid #0D1B2A; height: 22px; margin-bottom: 4px; }
.imzo-label { font-size: 9px; color: #9AC080; }
.muhr { width: 60px; height: 60px; border-radius: 50%; border: 2px solid #D8EDD0; display: flex; align-items: center; justify-content: center; font-size: 9px; color: #9AC080; margin: 0 auto; }
.qr-burchak { position: absolute; top: 6mm; right: 8mm; text-align: center; }
.qr-burchak img { width: 80px; height: 80px; }
.qr-burchak div { font-size: 8px; color: #607080; margin-top: 2px; }
"""


def _nakladnoy_nusxa_html(m: dict, sana: str, nusxa_nomi: str, qr_base64: str,
                           sahifa_uzilishi: bool = False) -> str:
    """Bitta Nakladnoy nusxasining (Zavod/Shofyor/Ohrana) HTML qismini
    quradi - ilova UI'sidagi karta-asosidagi ko'rinishga mos (brend
    yashili #0F6E56), 3 nusxa uchun ham AYNAN shu funksiya qayta
    ishlatiladi (faqat nusxa_nomi farqlanadi). [sahifa_uzilishi]=True
    bo'lsa, bu nusxa YANGI PDF sahifasidan boshlanadi (1-nusxada FALSE,
    2- va 3-nusxada TRUE bo'lishi kerak)."""
    uzilish_uslub = ' style="page-break-before: always;"' if sahifa_uzilishi else ""
    def arava_qatori(n):
        a = m["aravalar"].get(n)
        if not a or not a.get("tara"):
            return ""
        tara = a.get("tara") or 0
        brutto = a.get("brutto") or 0
        netto = a.get("netto")
        if netto is None:
            netto = brutto - tara
        kond = a.get("konditsion")
        kond_str = str(round(kond)) if kond is not None else "—"
        return (f"<tr><td>{n}-арава</td><td>{round(tara)}</td><td>{round(brutto)}</td>"
                f"<td>{round(netto)}</td><td>{kond_str}</td></tr>")

    aravalar_qatorlari = "".join(arava_qatori(n) for n in (1, 2, 3))

    jami_tara = sum((m["aravalar"].get(n) or {}).get("tara") or 0 for n in (1, 2, 3))
    jami_brutto = sum((m["aravalar"].get(n) or {}).get("brutto") or 0 for n in (1, 2, 3))
    jami_netto = jami_brutto - jami_tara
    jami_konditsion = sum((m["aravalar"].get(n) or {}).get("konditsion") or 0 for n in (1, 2, 3))

    namlik = m["namlik"] if m["namlik"] != "" else "—"
    ifloslik = m["ifloslik"] if m["ifloslik"] != "" else "—"
    qr_html = (
        f'<div class="qr-burchak"><img src="data:image/png;base64,{qr_base64}"/>'
        f'<div>Onlayn ko&#39;rish</div></div>'
    ) if qr_base64 else ""

    return f"""
<div class="sahifa"{uzilish_uslub}>
  {qr_html}
  <div class="nusxa-badge">{nusxa_nomi}</div>
  <div class="sarlavha">ТОВАР ТРАНСПОРТ НАКЛАДНОЙ № {m['raqam']}</div>
  <div class="subtitle">Ишлаб чиқаришдан қабул қилинган маҳсулотларни ташиш учун &nbsp;·&nbsp; {sana} &nbsp;·&nbsp; {m['mashina_turi']} {m['mashina_raqami']}</div>

  <div class="karta">
    <div class="maydon"><span class="label">Юк жўнатувчи</span><span class="qiymat">"Ҳазорасп текстил" МЧЖга қарашли пахта тозалаш заводи</span></div>
    <div class="maydon"><span class="label">Юк олувчи</span><span class="qiymat">{m['firma']}</span></div>
  </div>

  <div class="karta">
    <div class="info-grid">
      <div><span class="label">Тикет №</span><br><span class="qiymat">{m['tiket_raqam'] or '—'}</span></div>
      <div><span class="label">Туда №</span><br><span class="qiymat">{m['tuda_raqam'] or '—'}</span></div>
      <div><span class="label">Класс</span><br><span class="qiymat">{m['klass'] or '—'}</span></div>
      <div><span class="label">Селексия нави</span><br><span class="qiymat">{m['seleksiya_navi'] or '—'}</span></div>
      <div><span class="label">Терим тури</span><br><span class="qiymat">{m['terim_turi'] or '—'}</span></div>
      <div><span class="label">Намлик %</span><br><span class="qiymat">{namlik}</span></div>
      <div><span class="label">Ифлослик %</span><br><span class="qiymat">{ifloslik}</span></div>
      <div><span class="label">Шофёр</span><br><span class="qiymat">{m['shofyor']}</span></div>
    </div>
  </div>

  <table>
    <tr>
      <th>Юкнинг номи</th>
      <th>Тара (Урама), кг</th>
      <th>Брутто (Урама б/н), кг</th>
      <th>Нетто (Соф), кг</th>
      <th>Кондицион вазн, кг</th>
    </tr>
    {aravalar_qatorlari}
    <tr class="jami">
      <td>Жами:</td>
      <td>{round(jami_tara)}</td>
      <td>{round(jami_brutto)}</td>
      <td>{round(jami_netto)}</td>
      <td>{round(jami_konditsion)}</td>
    </tr>
  </table>

  <div class="dostaverka-box">
    <span><b>Доставерна № {m['dostaverka']}</b></span>
    <span>Муддат: {m['dostaverka_vaqt']}</span>
  </div>

  <div class="karta">
    <div class="maydon"><span class="label">Қабул қилди</span><span class="qiymat">{m['qabul_qildi'] or '—'} ___________</span></div>
    <div class="maydon"><span class="label">Юк олинди</span><span class="qiymat">{m['yuk_olindi'] or '—'} ___________</span></div>
  </div>

  <div class="imzo-grid">
    <div><div class="imzo-line"></div>Раҳбар<div class="imzo-label">ИМЗО</div></div>
    <div><div class="imzo-line"></div>Шофёр<div class="imzo-label">ИМЗО</div></div>
    <div><div class="imzo-line"></div>Юк олиб кетувчи<div class="imzo-label">ИМЗО</div></div>
    <div><div class="imzo-line"></div>Тарозибон<div class="imzo-label">ИМЗО</div></div>
    <div class="muhr">М.Ў.</div>
  </div>
</div>
"""


@app.post("/nakladnoy/saqlash")
def nakladnoy_saqlash(data: dict, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    try:
        hujjat_id = data.get("hujjat_id")
        if not hujjat_id:
            raise HTTPException(status_code=400, detail="hujjat_id kerak")

        m = nakladnoy_uchun_malumot(db, hujjat_id)
        if not m:
            raise HTTPException(status_code=404, detail=f"Hujjat topilmadi (id={hujjat_id})")

        hujjat_obj = db.query(Hujjat).filter(Hujjat.id == hujjat_id).first()
        if not hujjat_obj.nakladnoy_token:
            import secrets
            hujjat_obj.nakladnoy_token = secrets.token_urlsafe(24)
            db.commit()
            db.refresh(hujjat_obj)

        qr_base64 = ""
        try:
            import qrcode
            import io
            import base64
            korish_url = f"{SERVER_ASOSIY_URL}/nakladnoy-korish/{hujjat_obj.nakladnoy_token}"
            qr_rasm = qrcode.make(korish_url)
            buffer = io.BytesIO()
            qr_rasm.save(buffer, format="PNG")
            qr_base64 = base64.b64encode(buffer.getvalue()).decode("ascii")
        except Exception as qr_xato:
            print(f"QR kod generatsiya qilinmadi: {qr_xato}")

        sana = xavfsiz_sana(data.get("sana") or datetime.now().strftime("%Y-%m-%d"))

        # 3 nusxa - Zavod/Shofyor/Ohrana - AYNAN bir xil kontent, faqat
        # sarlavha nomi farqlanadi. QR kod har uchalasida ham bor (barcha
        # jismoniy nusxa mustaqil tekshirilishi mumkin bo'lishi uchun).
        html_content = f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8"><style>{_NAKLADNOY_STIL}</style></head>
<body>
{_nakladnoy_nusxa_html(m, sana, "ЗАВОД НУСХАСИ", qr_base64)}
{_nakladnoy_nusxa_html(m, sana, "ШОФЁР НУСХАСИ", qr_base64, sahifa_uzilishi=True)}
{_nakladnoy_nusxa_html(m, sana, "ОХРАНА НУСХАСИ", qr_base64, sahifa_uzilishi=True)}
</body></html>"""

        raqam_papka = xavfsiz_papka_nomi(
            (m["mashina_raqami"] or "noma_lum").replace(" ", "_"), "noma_lum")
        # {yil-oy}/{sana} - papkalar oy bo'yicha ham guruhlanadi (masalan
        # "2026-08/2026-08-01/"), shu bilan bitta mahsulot papkasi
        # ichida yiliga minglab kunlik papka to'planib qolmaydi. Excel
        # jurnali (excel_qatorga_yoz) va qo'lda eksport BUNGA kirmaydi -
        # ular ataylab faqat mahsulot darajasida (sana segmentisiz)
        # qoladi.
        papka = Path(f"C:/RASMLAR/{xavfsiz_papka_nomi(m['mahsulot_nomi'])}/{sana[:7]}/{sana}/{raqam_papka}")
        papka.mkdir(parents=True, exist_ok=True)

        html_fayl = papka / "nakladnoy.html"
        with open(html_fayl, "w", encoding="utf-8") as f:
            f.write(html_content)

        pdf_fayl = papka / "nakladnoy.pdf"
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            browser = p.chromium.launch()
            # Sahifa kengligi A4 landscape'ga teng (297mm). Balandlik esa
            # BITTA nusxaning haqiqiy kontent balandligiga moslab
            # hisoblanadi (uchala nusxa ham bir xil tuzilishga ega, shu
            # sabab birinchisini o'lchash yetarli) - shu o'lcham PDF
            # sahifa balandligi qilib beriladi, natijada har bir
            # `.sahifa` (CSS page-break-before bilan) aynan bitta PDF
            # sahifasiga to'g'ri keladi - 3 ta ALOHIDA sahifa (nusxa)
            # hosil bo'ladi (bitta uzun sahifa emas).
            kenglik_mm = 297
            kenglik_px = round(kenglik_mm * 96 / 25.4)
            page = browser.new_page(
                viewport={"width": kenglik_px, "height": 4800})
            page.goto(html_fayl.absolute().as_uri())
            kontent_px = page.evaluate(
                "document.querySelector('.sahifa').scrollHeight")
            balandlik_mm = kontent_px * 25.4 / 96 + 8
            page.pdf(path=str(pdf_fayl), width=f"{kenglik_mm}mm",
                     height=f"{balandlik_mm}mm", print_background=True,
                     margin={"top": "0mm", "bottom": "0mm",
                             "left": "0mm", "right": "0mm"})
            browser.close()

        # Fayl diskka (arxiv uchun, C:/RASMLAR/...) saqlanishda qoladi -
        # BUNDAN TASHQARI endi PDF baytlarining o'zi ham javob sifatida
        # qaytariladi, shunda frontend uni to'g'ridan-to'g'ri yuklab olishi
        # mumkin (Excel eksportidagi kabi - blob + anchor + download,
        # yangi oyna/tab OCHILMAYDI). Diskka yozish bilan bir vaqtda ishlagan
        # 3 ta chaqiruvchi (ekran tugmasi, offline-sinxronizatsiya, tortish
        # yakunlanganda avtomatik arxivlash) javob tanasini tekshirmaydi -
        # faqat status-kodga qaraydi, shu sabab bu o'zgarish ularga
        # ta'sir qilmaydi.
        pdf_baytlar = pdf_fayl.read_bytes()
        fayl_nomi = f"Nakladnoy_{m['raqam'].replace('/', '-')}.pdf"
        return Response(
            content=pdf_baytlar,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{fayl_nomi}"'},
        )
    except HTTPException:
        raise
    except Exception as e:
        print(f"XATO: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/nakladnoy-korish/{token}", response_class=HTMLResponse)
def nakladnoy_korish(token: str, db: Session = Depends(get_db)):
    """QR kod orqali ochiladigan, LOGIN TALAB QILMAYDIGAN ochiq sahifa.

    DIQQAT - xavfsizlik: bu yerda ATAYLAB hujjat_id emas, tasodifiy
    nakladnoy_token bo'yicha qidiriladi (Depends(get_current_user) YO'Q,
    lekin shu tufayli ID'larni ketma-ket sinab boshqa hujjatlarni ko'rish
    imkonsiz bo'ladi). Faqat GET - hech qanday yozish/o'zgartirish yo'q.
    """
    hujjat = db.query(Hujjat).filter(Hujjat.nakladnoy_token == token).first()
    if not hujjat:
        return HTMLResponse(
            content="<html><body style='font-family:Arial;text-align:center;padding:40px'>"
                    "<h2>Hujjat topilmadi</h2>"
                    "<p>Havola noto'g'ri yoki muddati o'tgan.</p></body></html>",
            status_code=404,
        )

    m = nakladnoy_uchun_malumot(db, hujjat.id)

    def arava_qatori(n):
        a = m["aravalar"].get(n)
        if not a or not a.get("tara"):
            return ""
        tara = a.get("tara") or 0
        brutto = a.get("brutto") or 0
        netto = a.get("netto")
        if netto is None:
            netto = brutto - tara
        kond = a.get("konditsion")
        kond_str = str(round(kond)) if kond is not None else "-"
        return (f"<tr><td>{n}-arava</td><td>{round(tara)}</td>"
                f"<td>{round(brutto)}</td><td>{round(netto)}</td><td>{kond_str}</td></tr>")

    aravalar_qatorlari = "".join(arava_qatori(n) for n in (1, 2, 3))

    jami_tara = sum((m["aravalar"].get(n) or {}).get("tara") or 0 for n in (1, 2, 3))
    jami_brutto = sum((m["aravalar"].get(n) or {}).get("brutto") or 0 for n in (1, 2, 3))
    jami_netto = jami_brutto - jami_tara
    jami_konditsion = sum((m["aravalar"].get(n) or {}).get("konditsion") or 0 for n in (1, 2, 3))

    namlik = m["namlik"] if m["namlik"] != "" else "—"
    ifloslik = m["ifloslik"] if m["ifloslik"] != "" else "—"

    html_content = f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Nakladnoy {m['raqam']}</title>
<style>
body {{ font-family: Arial, sans-serif; font-size: 16px; margin: 0; padding: 16px; background: #F5F7FA; color: #0D1B2A; }}
.karta {{ background: white; border-radius: 12px; padding: 16px; max-width: 480px; margin: 0 auto 12px auto; box-shadow: 0 1px 4px rgba(0,0,0,0.1); }}
h2 {{ font-size: 18px; margin: 0 0 4px 0; color: #1A4A08; }}
h3 {{ font-size: 13px; margin: 0 0 12px 0; color: #607080; font-weight: normal; }}
.maydon {{ display: flex; justify-content: space-between; padding: 6px 0; border-bottom: 1px solid #EEF1F4; font-size: 14px; }}
.maydon:last-child {{ border-bottom: none; }}
.label {{ color: #607080; }}
.qiymat {{ font-weight: 600; text-align: right; }}
table {{ width: 100%; border-collapse: collapse; margin-top: 8px; font-size: 13px; }}
th, td {{ border: 1px solid #E0E4E8; padding: 8px 6px; text-align: center; }}
th {{ background: #1A4A08; color: white; font-weight: 600; }}
.jami {{ font-weight: bold; background: #F0F4F0; }}
.vaqt-badge {{ display: inline-block; padding: 4px 10px; border-radius: 8px; font-size: 12px; margin-right: 6px; }}
.kirgan {{ background: #E3F2FD; color: #1565C0; }}
.chiqqan {{ background: #E8F5E9; color: #2E7D32; }}
</style></head>
<body>
<div class="karta">
  <h2>Nakladnoy № {m['raqam']}</h2>
  <h3>{m['mashina_turi']} · {m['mashina_raqami']} · {m['sana']}</h3>
  <div>
    <span class="vaqt-badge kirgan">Kirgan: {m['kirgan_vaqt'] or '—'}</span>
    <span class="vaqt-badge chiqqan">Chiqqan: {m['chiqqan_vaqt'] or '—'}</span>
  </div>
</div>
<div class="karta">
  <div class="maydon"><span class="label">Shofyor</span><span class="qiymat">{m['shofyor']}</span></div>
  <div class="maydon"><span class="label">Firma</span><span class="qiymat">{m['firma']}</span></div>
  <div class="maydon"><span class="label">Mahsulot</span><span class="qiymat">{m['mahsulot_nomi']}</span></div>
  <div class="maydon"><span class="label">Tiket №</span><span class="qiymat">{m['tiket_raqam'] or '—'}</span></div>
  <div class="maydon"><span class="label">Tuda №</span><span class="qiymat">{m['tuda_raqam'] or '—'}</span></div>
  <div class="maydon"><span class="label">Klass</span><span class="qiymat">{m['klass'] or '—'}</span></div>
  <div class="maydon"><span class="label">Sinf</span><span class="qiymat">{m['sinf'] or '—'}</span></div>
  <div class="maydon"><span class="label">Seleksiya navi</span><span class="qiymat">{m['seleksiya_navi'] or '—'}</span></div>
  <div class="maydon"><span class="label">Terim turi</span><span class="qiymat">{m['terim_turi'] or '—'}</span></div>
  <div class="maydon"><span class="label">Namlik %</span><span class="qiymat">{namlik}</span></div>
  <div class="maydon"><span class="label">Ifloslik %</span><span class="qiymat">{ifloslik}</span></div>
</div>
<div class="karta">
  <table>
    <tr><th>Arava</th><th>Tara</th><th>Brutto</th><th>Netto</th><th>Kond.</th></tr>
    {aravalar_qatorlari}
    <tr class="jami"><td>Jami</td><td>{round(jami_tara)}</td><td>{round(jami_brutto)}</td><td>{round(jami_netto)}</td><td>{round(jami_konditsion)}</td></tr>
  </table>
</div>
</body></html>"""
    return HTMLResponse(content=html_content)


# ============ KAMERA ============
import concurrent.futures

def _kichik_rasm_base64(rasm_baytlari: bytes) -> str:
    """Operator ekranida DARHOL ko'rsatish uchun kichik, siqilgan JPEG
    nusxa yaratadi (asl rasm hajmidan qat'i nazar, ~400px kenglik,
    sifat ~55%) - faqat vaqtinchalik vizual tasdiqlash uchun, hech
    qayerga saqlanmaydi. Kichik hajm tarmoq orqali tez kelishi va
    Flutter tomonida dekodlash ekранni "qotirmasligi" uchun muhim -
    asl (to'liq sifatli) rasm bundan mustaqil, alohida diskka
    saqlanadi (qarang: bir_kameradan_rasm_ol())."""
    import base64
    from PIL import Image
    rasm = Image.open(io.BytesIO(rasm_baytlari))
    if rasm.mode != "RGB":
        rasm = rasm.convert("RGB")
    kenglik_maksimal = 400
    if rasm.width > kenglik_maksimal:
        nisbat = kenglik_maksimal / rasm.width
        rasm = rasm.resize((kenglik_maksimal, round(rasm.height * nisbat)))
    buffer = io.BytesIO()
    rasm.save(buffer, format="JPEG", quality=55)
    return base64.b64encode(buffer.getvalue()).decode("ascii")


def bir_kameradan_rasm_ol(cam_ip, fayl_yol):
    try:
        url = f"http://{cam_ip}/ISAPI/Streaming/channels/101/picture"
        response = req.get(
            url,
            auth=req.auth.HTTPDigestAuth(KAMERA_LOGIN, KAMERA_PAROL),
            timeout=5
        )
        if response.status_code == 200:
            with open(fayl_yol, "wb") as f:
                f.write(response.content)
            natija = {"status": "ok", "fayl": str(fayl_yol)}
            # Kichraytirish muvaffaqiyatsiz bo'lsa ham (masalan kutilmagan
            # rasm formati) asosiy natija ("fayl saqlandi") o'zgarmasligi
            # kerak - faqat operator ekranida ko'rsatib bo'lmaydi, xolos.
            try:
                natija["rasm_base64"] = _kichik_rasm_base64(response.content)
            except Exception:
                pass
            return natija
        else:
            xabar = f"Kamera {cam_ip} javob bermadi"
            tizim_xatosini_saqla("kamera", xabar)
            return {"status": "error", "message": xabar}
    except Exception as e:
        tizim_xatosini_saqla("kamera", str(e))
        return {"status": "error", "message": str(e)}


# Kamera Telegram ogohlantirishi: har bir arava saqlanganda kamera
# chaqirilgani uchun (kuniga o'nlab/yuzlab marta) HAR bir xatoda xabar
# yuborilsa Telegram guruh spam bilan to'lib ketadi. Shu sabab har
# kamera (IP) uchun ALOHIDA ketma-ket xatolar hisoblanadi - faqat
# KAMERA_XATO_CHEGARA marta ketma-ket muvaffaqiyatsiz bo'lsa bitta
# ogohlantirish yuboriladi, keyingi xatolarda esa (tuzalmaguncha)
# qayta yuborilmaydi. Holat xotirada saqlanadi (DB shart emas) -
# server qayta ishga tushirilsa hisoblagich nolga tushadi, bu
# zararsiz (faqat qayta 3 marta xato kerak bo'ladi, xolos).
KAMERA_XATO_CHEGARA = 3

_kamera_holati = {
    KAMERA_1_IP: {"ketma_ket": 0, "ogohlantirilgan": False},
    KAMERA_2_IP: {"ketma_ket": 0, "ogohlantirilgan": False},
}
# _kamera_holati sinxron route handler'dan (FastAPI thread pool'da
# PARALLEL ishlaydi) o'zgartiriladi - shu sabab boshqa global holatlar
# (_tarozi_holat, _excel_qulflari) kabi Lock bilan himoyalanadi, aks
# holda ikki operator stansiyasi bir vaqtda rasm yuborsa hisoblagich
# noto'g'ri sanashi yoki ogohlantirish ikki marta yuborilishi mumkin edi.
_kamera_holati_qulf = threading.Lock()


def kamera_xatosi_ogohlantirish(cam_nomi: str, cam_ip: str, natija: dict,
                                 mashina_raqami: str, mahsulot_nomi: str, tur: str):
    # Telegram so'rovi (tarmoq I/O) LOCK OSTIDA YUBORILMAYDI - faqat
    # holatni o'qish/yangilash va "yuborish kerakmi" qarorini qulf ichida
    # qilamiz, xabarning o'zini qulfdan chiqqach yuboramiz.
    yuboriladigan_matn = None
    with _kamera_holati_qulf:
        holat = _kamera_holati.get(cam_ip)
        if holat is None:
            holat = {"ketma_ket": 0, "ogohlantirilgan": False}
            _kamera_holati[cam_ip] = holat

        if natija["status"] == "error":
            holat["ketma_ket"] += 1
            if holat["ketma_ket"] >= KAMERA_XATO_CHEGARA and not holat["ogohlantirilgan"]:
                holat["ogohlantirilgan"] = True
                vaqt = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                # html.escape(): natija["message"] xom Python exception matni
                # bo'lishi mumkin (masalan "<HTTPConnection(...) at 0x...>"
                # kabi burchakli qavslar bilan) - parse_mode=HTML bunday
                # matnni noto'g'ri teg deb Telegram butun xabarni rad etadi
                # (400 Bad Request), aynan ogohlantirish kerak bo'lgan paytda
                # yetib bormay qolishi mumkin edi.
                yuboriladigan_matn = (
                    f"⚠️ <b>Kamera xatosi</b>\n\n"
                    f"Kamera: {html.escape(cam_nomi)} ({html.escape(cam_ip)})\n"
                    f"Mashina: {html.escape(mashina_raqami)}\n"
                    f"Mahsulot: {html.escape(mahsulot_nomi)} ({html.escape(tur)})\n"
                    f"Vaqt: {vaqt}\n"
                    f"Xato: {html.escape(natija['message'])}\n\n"
                    f"({holat['ketma_ket']} marta ketma-ket muvaffaqiyatsiz)"
                )
        else:
            if holat["ogohlantirilgan"]:
                yuboriladigan_matn = f"✅ Kamera {cam_nomi} ({cam_ip}) qayta ishlay boshladi."
            holat["ketma_ket"] = 0
            holat["ogohlantirilgan"] = False

    if yuboriladigan_matn:
        telegram_xabar_yuborish(yuboriladigan_matn)


@app.post("/kamera/rasm")
def rasm_ol(data: dict, current_user: dict = Depends(get_current_user)):
    # DIQQAT: bu qiymatlar operator ekranidagi JONLI matn maydonidan
    # (raqamiCtrl.text) keladi va oxirida bo'sh joy qolib ketishi
    # mumkin (masalan avtomatik to'ldirish/bosish natijasida) - .strip()
    # QILINMASA, xuddi shu mashina uchun keyinroq (Hujjat.mashina_raqami
    # orqali, allaqachon tozalangan holda) yaratiladigan nakladnoy
    # papkasidan BOSHQA papka hosil bo'lib qoladi (masalan "90_90_90"
    # va "90_90_90_" - ikkita alohida papka, rasmlar va nakladnoy
    # ajralib qoladi).
    mashina_raqami = data.get("mashina_raqami", "noma_lum").strip()
    mahsulot_nomi = data.get("mahsulot_nomi", "Chigit").strip()
    tur = data.get("tur", "tara")

    sana = datetime.now().strftime("%Y-%m-%d")
    vaqt = datetime.now().strftime("%H-%M-%S")
    raqam = xavfsiz_papka_nomi(mashina_raqami.replace(" ", "_"), "noma_lum")

    # {yil-oy}/{sana} - nakladnoydagi bilan bir xil tuzilish (qarang:
    # nakladnoy_saqlash) - shu ikkalasi ALOHIDA joyda hisoblansada,
    # bitta mahsulot+sana+mashina uchun AYNAN bir xil papkaga tushishi
    # SHART, aks holda rasmlar va nakladnoy ikkiga bo'linib qoladi.
    papka = Path(f"C:/RASMLAR/{xavfsiz_papka_nomi(mahsulot_nomi, 'Chigit')}/{sana[:7]}/{sana}/{raqam}")
    papka.mkdir(parents=True, exist_ok=True)

    fayl1 = papka / f"{tur}_cam1_{vaqt}.jpg"
    fayl2 = papka / f"{tur}_cam2_{vaqt}.jpg"

    with concurrent.futures.ThreadPoolExecutor(max_workers=2) as executor:
        future1 = executor.submit(bir_kameradan_rasm_ol, KAMERA_1_IP, fayl1)
        future2 = executor.submit(bir_kameradan_rasm_ol, KAMERA_2_IP, fayl2)
        natija1 = future1.result()
        natija2 = future2.result()

    kamera_xatosi_ogohlantirish("Kamera 1", KAMERA_1_IP, natija1, mashina_raqami, mahsulot_nomi, tur)
    kamera_xatosi_ogohlantirish("Kamera 2", KAMERA_2_IP, natija2, mashina_raqami, mahsulot_nomi, tur)

    # Ikkala kamera ham muvaffaqiyatsiz bo'lsa - bu haqiqiy "yuqori oqim"
    # (kamera qurilmasi) xatosi, frontend buni ANIQ ko'rishi va offline
    # qayta-urinish navbatiga QO'YMASLIGI kerak (qayta urinish kamera
    # holatini o'zgartirmaydi). Faqat bitta kamera ishlamasa - bu qisman
    # muvaffaqiyat (bitta rasm baribir olindi), 200 bilan qoladi.
    if natija1["status"] == "error" and natija2["status"] == "error":
        raise HTTPException(
            status_code=502,
            detail=f"Ikkala kamera ham javob bermadi: {natija1['message']} / {natija2['message']}",
        )

    return {
        "status": "ok",
        "vaqt": vaqt,
        "kamera1": natija1,
        "kamera2": natija2,
        "papka": str(papka)
    }

# ============ GRAFIK MA'LUMOTLAR ============

@app.get("/statistika/grafik/kunlik")
def grafik_kunlik(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    from datetime import date, timedelta
    kun_boshi = date.today() - timedelta(days=6)
    oxirgi_kun = date.today() + timedelta(days=1)

    qatorlar = db.query(
        cast(Hujjat.created_at, Date).label('kun'),
        Hujjat.mahsulot_id,
        func.count(func.distinct(Hujjat.id)).label('soni'),
    ).filter(
        Hujjat.created_at >= kun_boshi,
        Hujjat.created_at < oxirgi_kun
    ).group_by(cast(Hujjat.created_at, Date), Hujjat.mahsulot_id).all()

    mahsulot_dict = {}
    jami_dict = {}
    for row in qatorlar:
        kun_str = str(row.kun)
        mahsulot_dict[(kun_str, row.mahsulot_id)] = row.soni
        jami_dict[kun_str] = jami_dict.get(kun_str, 0) + row.soni

    natija = []
    for i in range(6, -1, -1):
        kun = date.today() - timedelta(days=i)
        kun_str = str(kun)
        natija.append({
            "kun": kun_str,
            "chigit": mahsulot_dict.get((kun_str, 1), 0),
            "chiganoq": mahsulot_dict.get((kun_str, 2), 0),
            "pochog": mahsulot_dict.get((kun_str, 3), 0),
            "jami": jami_dict.get(kun_str, 0),
        })
    return natija

@app.get("/statistika/grafik/haftalik")
def grafik_haftalik(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    from datetime import date, timedelta
    bugun = date.today()
    joriy_hafta_boshi = bugun - timedelta(days=bugun.weekday())
    hafta_boshi_8 = joriy_hafta_boshi - timedelta(weeks=7)
    oxirgi_chegara = bugun + timedelta(days=1)

    qatorlar = db.query(
        func.date_trunc('week', Hujjat.created_at).label('hafta'),
        Hujjat.mahsulot_id,
        func.count(func.distinct(Hujjat.id)).label('soni'),
    ).filter(
        Hujjat.created_at >= hafta_boshi_8,
        Hujjat.created_at < oxirgi_chegara
    ).group_by(func.date_trunc('week', Hujjat.created_at), Hujjat.mahsulot_id).all()

    mahsulot_dict = {}
    jami_dict = {}
    for row in qatorlar:
        hafta_str = str(row.hafta.date())
        mahsulot_dict[(hafta_str, row.mahsulot_id)] = row.soni
        jami_dict[hafta_str] = jami_dict.get(hafta_str, 0) + row.soni

    natija = []
    for i in range(7, -1, -1):
        hafta = joriy_hafta_boshi - timedelta(weeks=i)
        hafta_str = str(hafta)
        natija.append({
            "hafta_boshi": hafta_str,
            "chigit": mahsulot_dict.get((hafta_str, 1), 0),
            "chiganoq": mahsulot_dict.get((hafta_str, 2), 0),
            "pochog": mahsulot_dict.get((hafta_str, 3), 0),
            "jami": jami_dict.get(hafta_str, 0),
        })
    return natija

@app.get("/statistika/grafik/oylik")
def grafik_oylik(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    from datetime import date, timedelta
    bugun = date.today()
    oy_boshi = bugun.replace(day=1)
    oxirgi_kun = bugun + timedelta(days=1)

    qatorlar = db.query(
        cast(Hujjat.created_at, Date).label('kun'),
        Hujjat.mahsulot_id,
        func.count(func.distinct(Hujjat.id)).label('soni'),
    ).filter(
        Hujjat.created_at >= oy_boshi,
        Hujjat.created_at < oxirgi_kun
    ).group_by(cast(Hujjat.created_at, Date), Hujjat.mahsulot_id).all()

    mahsulot_dict = {}
    jami_dict = {}
    for row in qatorlar:
        kun_str = str(row.kun)
        mahsulot_dict[(kun_str, row.mahsulot_id)] = row.soni
        jami_dict[kun_str] = jami_dict.get(kun_str, 0) + row.soni

    natija = []
    kun = oy_boshi
    while kun <= bugun:
        kun_str = str(kun)
        natija.append({
            "kun": kun_str,
            "chigit": mahsulot_dict.get((kun_str, 1), 0),
            "chiganoq": mahsulot_dict.get((kun_str, 2), 0),
            "pochog": mahsulot_dict.get((kun_str, 3), 0),
            "jami": jami_dict.get(kun_str, 0),
        })
        kun = kun + timedelta(days=1)
    return natija

@app.get("/statistika/grafik/mavsum")
def grafik_mavsum(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    from datetime import date
    bugun = date.today()
    if bugun.month >= 8:
        mavsum_boshi = date(bugun.year, 8, 1)
    else:
        mavsum_boshi = date(bugun.year - 1, 8, 1)
    oxirgi_oy = date(bugun.year + (bugun.month // 12), (bugun.month % 12) + 1, 1)

    qatorlar = db.query(
        func.date_trunc('month', Hujjat.created_at).label('oy'),
        Hujjat.mahsulot_id,
        func.count(func.distinct(Hujjat.id)).label('soni'),
    ).filter(
        Hujjat.created_at >= mavsum_boshi,
        Hujjat.created_at < oxirgi_oy
    ).group_by(func.date_trunc('month', Hujjat.created_at), Hujjat.mahsulot_id).all()

    mahsulot_dict = {}
    jami_dict = {}
    for row in qatorlar:
        oy_str = row.oy.strftime("%Y-%m")
        mahsulot_dict[(oy_str, row.mahsulot_id)] = row.soni
        jami_dict[oy_str] = jami_dict.get(oy_str, 0) + row.soni

    natija = []
    oy = mavsum_boshi
    while oy <= bugun:
        oy_str = f"{oy.year}-{oy.month:02d}"
        natija.append({
            "oy": oy_str,
            "chigit": mahsulot_dict.get((oy_str, 1), 0),
            "chiganoq": mahsulot_dict.get((oy_str, 2), 0),
            "pochog": mahsulot_dict.get((oy_str, 3), 0),
            "jami": jami_dict.get(oy_str, 0),
        })
        oy = date(oy.year + (oy.month // 12), (oy.month % 12) + 1, 1)
    return natija

# ============ GRAFIK-DETAL (mahsulot nomi + davr bo'yicha) ============

def _mahsulot_id_topish(db: Session, nom: str) -> int:
    mahsulot = db.query(Mahsulot).filter(func.lower(Mahsulot.nom) == nom.lower()).first()
    if not mahsulot:
        raise HTTPException(status_code=404, detail=f"'{nom}' nomli mahsulot topilmadi!")
    return mahsulot.id

@app.get("/statistika/grafik-detal/kunlik")
def grafik_detal_kunlik(mahsulot: str, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    from datetime import date, timedelta
    mahsulot_id = _mahsulot_id_topish(db, mahsulot)

    bugun = date.today()
    ertaga = bugun + timedelta(days=1)

    soat_ustuni = func.date_part('hour', Hujjat.created_at)
    qatorlar = db.query(
        soat_ustuni.label('soat'),
        func.count(func.distinct(Hujjat.id)).label('soni'),
        func.coalesce(func.sum(Olchov.netto), 0).label('jami_netto'),
    ).outerjoin(
        Olchov, Olchov.hujjat_id == Hujjat.id
    ).filter(
        Hujjat.mahsulot_id == mahsulot_id,
        Hujjat.created_at >= bugun,
        Hujjat.created_at < ertaga,
    ).group_by(soat_ustuni).all()

    soat_dict = {int(row.soat): (row.soni, row.jami_netto) for row in qatorlar}

    natija = []
    for soat in range(24):
        soni, jami_netto = soat_dict.get(soat, (0, 0))
        natija.append({
            "soat": soat,
            "soni": soni,
            "tonnaj": round(jami_netto / 1000, 2),
        })
    return natija

@app.get("/statistika/grafik-detal/haftalik")
def grafik_detal_haftalik(mahsulot: str, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    from datetime import date, timedelta
    mahsulot_id = _mahsulot_id_topish(db, mahsulot)

    bugun = date.today()
    hafta_boshi = bugun - timedelta(days=bugun.weekday())
    keyingi_hafta = hafta_boshi + timedelta(days=7)

    kun_ustuni = func.date_part('isodow', Hujjat.created_at)
    qatorlar = db.query(
        kun_ustuni.label('kun'),
        func.count(func.distinct(Hujjat.id)).label('soni'),
        func.coalesce(func.sum(Olchov.netto), 0).label('jami_netto'),
    ).outerjoin(
        Olchov, Olchov.hujjat_id == Hujjat.id
    ).filter(
        Hujjat.mahsulot_id == mahsulot_id,
        Hujjat.created_at >= hafta_boshi,
        Hujjat.created_at < keyingi_hafta,
    ).group_by(kun_ustuni).all()

    kun_dict = {int(row.kun): (row.soni, row.jami_netto) for row in qatorlar}

    natija = []
    for kun_raqami in range(1, 8):
        soni, jami_netto = kun_dict.get(kun_raqami, (0, 0))
        natija.append({
            "kun_raqami": kun_raqami,
            "soni": soni,
            "tonnaj": round(jami_netto / 1000, 2),
        })
    return natija

@app.get("/statistika/grafik-detal/oylik")
def grafik_detal_oylik(mahsulot: str, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    from datetime import date
    mahsulot_id = _mahsulot_id_topish(db, mahsulot)

    bugun = date.today()
    oy_boshi = bugun.replace(day=1)
    if oy_boshi.month == 12:
        keyingi_oy = date(oy_boshi.year + 1, 1, 1)
    else:
        keyingi_oy = date(oy_boshi.year, oy_boshi.month + 1, 1)
    kunlar_soni = (keyingi_oy - oy_boshi).days

    kun_ustuni = func.date_part('day', Hujjat.created_at)
    qatorlar = db.query(
        kun_ustuni.label('kun'),
        func.count(func.distinct(Hujjat.id)).label('soni'),
        func.coalesce(func.sum(Olchov.netto), 0).label('jami_netto'),
    ).outerjoin(
        Olchov, Olchov.hujjat_id == Hujjat.id
    ).filter(
        Hujjat.mahsulot_id == mahsulot_id,
        Hujjat.created_at >= oy_boshi,
        Hujjat.created_at < keyingi_oy,
    ).group_by(kun_ustuni).all()

    kun_dict = {int(row.kun): (row.soni, row.jami_netto) for row in qatorlar}

    natija = []
    for kun in range(1, kunlar_soni + 1):
        soni, jami_netto = kun_dict.get(kun, (0, 0))
        natija.append({
            "kun": kun,
            "soni": soni,
            "tonnaj": round(jami_netto / 1000, 2),
        })
    return natija

@app.get("/statistika/grafik-detal/mavsum")
def grafik_detal_mavsum(mahsulot: str, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    from datetime import date
    mahsulot_id = _mahsulot_id_topish(db, mahsulot)

    bugun = date.today()
    if bugun.month >= 8:
        mavsum_boshi = date(bugun.year, 8, 1)
    else:
        mavsum_boshi = date(bugun.year - 1, 8, 1)
    mavsum_oxiri = date(mavsum_boshi.year + 1, 8, 1)

    oy_ustuni = func.date_trunc('month', Hujjat.created_at)
    qatorlar = db.query(
        oy_ustuni.label('oy'),
        func.count(func.distinct(Hujjat.id)).label('soni'),
        func.coalesce(func.sum(Olchov.netto), 0).label('jami_netto'),
    ).outerjoin(
        Olchov, Olchov.hujjat_id == Hujjat.id
    ).filter(
        Hujjat.mahsulot_id == mahsulot_id,
        Hujjat.created_at >= mavsum_boshi,
        Hujjat.created_at < mavsum_oxiri,
    ).group_by(oy_ustuni).all()

    oy_dict = {(row.oy.year, row.oy.month): (row.soni, row.jami_netto) for row in qatorlar}

    natija = []
    oy = mavsum_boshi
    for _ in range(12):
        soni, jami_netto = oy_dict.get((oy.year, oy.month), (0, 0))
        natija.append({
            "oy": oy.month,
            "yil": oy.year,
            "soni": soni,
            "tonnaj": round(jami_netto / 1000, 2),
        })
        if oy.month == 12:
            oy = date(oy.year + 1, 1, 1)
        else:
            oy = date(oy.year, oy.month + 1, 1)
    return natija
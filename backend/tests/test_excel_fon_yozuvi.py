"""
excel_qatorga_yoz_fon() (Excel jurnalini FON vazifasi sifatida yozish -
operator "Saqlash" javobini kutib turmasligi uchun) uchun ikkita
maqsadli test:

  1. Mazmun bir xilligi - fon orqali yozilgan HAQIQIY faylning mazmuni
     kutilgan qiymatlar bilan mos ekanini, faylning o'zini o'qib
     tekshiradi (mavjud testlar excel_qatorga_yoz()ni doim monkeypatch
     bilan "o'chirib qo'yadi" - haqiqiy fayl mazmunini hech biri
     tekshirmaydi, bu yerda birinchi marta tekshiriladi).
  2. Tartib/musobaqa xavfsizligi - bir xil mahsulot uchun IKKITA hujjat
     deyarli bir vaqtda (ikki alohida thread'da, ikkita mustaqil fon
     chaqiruvi bilan) yozilsa, ikkalasi ham natijaviy faylda - biri
     ikkinchisining ustidan yozib, yo'qotib qo'ymasdan - birga
     qolishini tekshiradi.

DIQQAT - bu fayl repo'dagi boshqa testlardan farqli, standart `client`/
`db_session` fixturalarini ISHLATMAYDI. Sababi: excel_qatorga_yoz_fon()
o'zining ALOHIDA (yangi) DB ulanishini ochadi - bu ulanish standart
`db_session` fixturasining SAVEPOINT-asosidagi (hech qachon haqiqiy
COMMIT qilinmaydigan, test oxirida ROLLBACK qilinadigan) tranzaksiyasini
KO'RA OLMAYDI (PostgreSQL'da committed bo'lmagan ma'lumot boshqa
ulanishga umuman ko'rinmaydi). Shu sabab bu yerda ma'lumotlar HAQIQATAN
(alohida SessionLocal() + real commit bilan) yoziladi va test oxirida
QO'LDA tozalanadi - xuddi audit 1-bosqichidagi (yuklama) benchmark
skriptlaridagi xavfsizlik naqshi bilan bir xil: soxta, aniq
ajratiladigan mahsulot nomi ishlatiladi (C:/RASMLAR/<nom>/... - haqiqiy
Chigit/Chiganoq fayllariga HECH QANDAY ta'sir qilmaydi), test oldidan
VA oxirida (muvaffaqiyatli yoki muvaffaqiyatsiz bo'lishidan qat'i
nazar) barcha DB qatorlari + disk papkasi butunlay o'chiriladi.
"""
import shutil
import threading
from datetime import datetime
from pathlib import Path

import openpyxl
import pytest

import main
from database import SessionLocal, engine
from models import Hujjat, HujjatHolati, Mahsulot, Olchov
from utils import konditsion_hisobla

MAHSULOT_NOMI = "AuditSinovMahsuloti"
RAQAM_PREFIKS = "AUDITSINOV-"


def _tozala():
    db = SessionLocal()
    try:
        eski = db.query(Hujjat).filter(Hujjat.raqam.like(f"{RAQAM_PREFIKS}%")).all()
        idlar = [h.id for h in eski]
        if idlar:
            db.query(Olchov).filter(Olchov.hujjat_id.in_(idlar)).delete(synchronize_session=False)
            db.query(Hujjat).filter(Hujjat.id.in_(idlar)).delete(synchronize_session=False)
        db.query(Mahsulot).filter(Mahsulot.nom == MAHSULOT_NOMI).delete(synchronize_session=False)
        db.commit()
    finally:
        db.close()
    papka = Path(f"C:/RASMLAR/{MAHSULOT_NOMI}")
    if papka.exists():
        shutil.rmtree(papka)


@pytest.fixture()
def sinov_mahsuloti():
    _tozala()  # oldingi uzilib qolgan urinishdan qoldiq bo'lsa, tozalab boshlanadi
    db = SessionLocal()
    try:
        mahsulot = Mahsulot(nom=MAHSULOT_NOMI, konditsiya_bor=True, is_active=True)
        db.add(mahsulot)
        db.commit()
        db.refresh(mahsulot)
        mahsulot_id = mahsulot.id
    finally:
        db.close()

    yield mahsulot_id

    _tozala()


def _hujjat_yarat(mahsulot_id, raqam_suffiks, tara, brutto, namlik=8.0, ifloslik=2.0):
    """Bitta TUGALLANDI holatidagi hujjat + o'lchov qatorini HAQIQATAN
    (real commit bilan) yaratadi va hujjat_id'ni qaytaradi."""
    db = SessionLocal()
    try:
        yil = datetime.now().year
        h = Hujjat(
            mahsulot_id=mahsulot_id,
            raqam=f"{RAQAM_PREFIKS}{yil}/{raqam_suffiks}",
            mashina_raqami=f"90 SINOV {raqam_suffiks}",
            shofyor="Sinov Shofyor", firma="Sinov MChJ",
            holat=HujjatHolati.TUGALLANDI,
            created_at=datetime.now(),
        )
        db.add(h)
        db.flush()
        netto = brutto - tara
        konditsion = konditsion_hisobla(netto, namlik, ifloslik)
        db.add(Olchov(hujjat_id=h.id, arava_raqam=1, tara=tara, brutto=brutto,
                       netto=netto, namlik=namlik, ifloslik=ifloslik, konditsion=konditsion))
        db.commit()
        return h.id, netto, konditsion
    finally:
        db.close()


def _fayl_yoli():
    yil = datetime.now().year
    return Path(f"C:/RASMLAR/{MAHSULOT_NOMI}/hisobot_{MAHSULOT_NOMI}_{yil}.xlsx")


def _qatorni_top(ws, mashina_raqami):
    """Berilgan mashina raqamiga mos qatorni topib, ustunlar ro'yxati
    sifatida qaytaradi (kun/oy guruhlash sarlavha qatorlari orasida
    ANIQ QAYERDA joylashganini bilish shart emas - shu bilan test
    guruhlash tafsilotlariga bog'liq bo'lmaydi, faqat "mening ma'lumotim
    to'g'ri yozildimi" ni tekshiradi)."""
    for row in ws.iter_rows(values_only=True):
        if mashina_raqami in row:
            return row
    return None


def test_fon_yozuvi_mazmuni_togri(sinov_mahsuloti):
    hujjat_id, kutilgan_netto, kutilgan_konditsion = _hujjat_yarat(
        sinov_mahsuloti, "001", tara=18000, brutto=25500, namlik=8.5, ifloslik=2.0)

    band_ulanishlar_oldin = engine.pool.checkedout()
    main.excel_qatorga_yoz_fon(hujjat_id)
    # excel_qatorga_yoz_fon() o'z DB ulanishini ochib-yopishi kerak -
    # chaqiruvdan keyin pool'da "unutilgan" band ulanish qolmasligi
    # kerak (resurs sizib chiqishining oldini olish, audit 2-bosqichi
    # bilan bog'liq).
    assert engine.pool.checkedout() == band_ulanishlar_oldin

    fayl = _fayl_yoli()
    assert fayl.exists(), f"Excel fayl yaratilmadi: {fayl}"

    wb = openpyxl.load_workbook(fayl)
    ws = wb.active
    qator = _qatorni_top(ws, "90 SINOV 001")
    assert qator is not None, "Yozilgan hujjatning qatori faylda topilmadi"

    # Ustunlar tartibi (konditsiya_bor=True mahsulot uchun):
    # tartib, raqam, mahsulot_nomi, sana, tara, brutto, netto, konditsion,
    # mashina_raqami, shofyor, firma, ...
    assert qator[4] == 18000  # tara
    assert qator[5] == 25500  # brutto
    assert qator[6] == round(kutilgan_netto)  # netto
    assert qator[7] == round(kutilgan_konditsion)  # konditsion
    assert qator[8] == "90 SINOV 001"  # mashina_raqami
    assert qator[9] == "Sinov Shofyor"
    assert qator[10] == "Sinov MChJ"


def test_fon_yozuvi_tartib_musobaqa_xavfsiz(sinov_mahsuloti):
    """Ikkita hujjat (bir xil mahsulot) deyarli bir vaqtda, ikki
    ALOHIDA thread'da fon vazifasi sifatida yozilsa - ikkalasi ham
    (qaysi biri oxirida tugashidan qat'i nazar) natijaviy faylda birga
    bo'lishi kerak. Bu excel_qatorga_yoz() har doim BAZADAN TO'LIQ
    QAYTA o'qishi (hech qachon eski/qisman "keshlangan" holatdan
    foydalanmasligi) + fayl bo'yicha threading.Lock borligi tufayli
    kafolatlanadi (qarang: main.py'dagi _excel_fayl_qulfi)."""
    hujjat_id_1, _, _ = _hujjat_yarat(sinov_mahsuloti, "101", tara=18000, brutto=24000)
    hujjat_id_2, _, _ = _hujjat_yarat(sinov_mahsuloti, "102", tara=18500, brutto=26000)

    boshlash_toʻsigʻi = threading.Barrier(2)
    xatolar = []

    def _fon_chaqiruv(hujjat_id):
        try:
            boshlash_toʻsigʻi.wait(timeout=5)  # ikkalasi ham AYNAN bir vaqtda boshlansin
            main.excel_qatorga_yoz_fon(hujjat_id)
        except Exception as e:  # pragma: no cover - faqat diagnostika uchun
            xatolar.append(e)

    t1 = threading.Thread(target=_fon_chaqiruv, args=(hujjat_id_1,))
    t2 = threading.Thread(target=_fon_chaqiruv, args=(hujjat_id_2,))
    t1.start()
    t2.start()
    t1.join(timeout=15)
    t2.join(timeout=15)

    assert not xatolar, f"Fon chaqiruvlarida kutilmagan xato: {xatolar}"

    fayl = _fayl_yoli()
    assert fayl.exists()
    wb = openpyxl.load_workbook(fayl)
    ws = wb.active

    qator1 = _qatorni_top(ws, "90 SINOV 101")
    qator2 = _qatorni_top(ws, "90 SINOV 102")
    assert qator1 is not None, "Birinchi hujjat qatori yo'qolib qoldi (musobaqa xatosi)"
    assert qator2 is not None, "Ikkinchi hujjat qatori yo'qolib qoldi (musobaqa xatosi)"
    assert qator1[4] == 18000 and qator1[5] == 24000
    assert qator2[4] == 18500 and qator2[5] == 26000
